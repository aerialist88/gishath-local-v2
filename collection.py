"""collection.py — your Moxfield collection, priced against Singapore stores.

Three pieces in one module:

  1. STORAGE — SQLite at state/collection.db. `items` mirrors the Moxfield
     CSV export (one row per name/set/number/foil/condition line); `prices`
     holds the latest match per item; `value_snapshots` records the running
     total once per completed pricing pass (collection-value-over-time).

  2. MATCHER — turns a /search result set into a per-item price with an
     honest confidence tier:
        exact — set matched AND the print's variant traits all appear in the
                listing text (and a plain print has no conflicting variant
                markers) AND foil status matches
        set   — set matched + foil matched, variant unconfirmed
        any   — cheapest listing of any print (foil-matching preferred)
        none  — no in-stock listing anywhere
     Print identity (set name + variant traits) comes from print_index.py;
     when that index has no entry (or isn't built yet) items cap out at "any".

  3. PRICER — a paced background worker. Groups unpriced/stale items by card
     name, POSTs batches of names to the app's own /search (reusing the full
     engine+Playwright pipeline, its error handling, and — for free — the
     price-history recorder), sleeps between batches so a thousands-card
     collection is a slow polite drip rather than a store-hammering flood.
     Pause/resume from the UI; progress survives restarts because every
     batch's results are committed to SQLite as they land.

Sizing honesty: at ~10 names/batch and ~30-45s per cycle, a 2,000-name
collection takes roughly 2 hours per full pass. Run it overnight alongside
whatever else; re-pricing only touches items older than REFRESH_AGE_DAYS.
"""
from __future__ import annotations

import csv
import io
import logging
import os
import sqlite3
import threading
import time
from datetime import date, datetime

import httpx

import print_index

log = logging.getLogger(__name__)

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(_BASE_DIR, "state", "collection.db")

APP_SEARCH_URL = "http://127.0.0.1:5003/search"
BATCH_NAMES = 10           # unique card names per /search call
BATCH_PAUSE_S = 10.0       # polite gap between batches
SEARCH_TIMEOUT_S = 120.0   # /search's own budget is 60s
REFRESH_AGE_DAYS = 7       # re-price items not checked within this window

_SCHEMA = """
CREATE TABLE IF NOT EXISTS items (
    id INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    set_code TEXT NOT NULL DEFAULT '',
    collector_number TEXT NOT NULL DEFAULT '',
    foil TEXT NOT NULL DEFAULT '',
    condition TEXT NOT NULL DEFAULT '',
    language TEXT NOT NULL DEFAULT '',
    count INTEGER NOT NULL DEFAULT 1,
    purchase_price REAL,
    UNIQUE(name, set_code, collector_number, foil, condition, language)
);
CREATE TABLE IF NOT EXISTS prices (
    item_id INTEGER PRIMARY KEY REFERENCES items(id) ON DELETE CASCADE,
    checked_at TEXT NOT NULL,
    tier TEXT NOT NULL,
    price REAL,
    store TEXT,
    url TEXT,
    listing TEXT
);
CREATE TABLE IF NOT EXISTS value_snapshots (
    day TEXT PRIMARY KEY,
    total_sgd REAL NOT NULL,
    priced INTEGER NOT NULL,
    items INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_items_name ON items (name);
"""


def _connect() -> sqlite3.Connection:
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=10.0)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(_SCHEMA)
    # Migration: foil_gap arrived after the first release of this table.
    try:
        conn.execute("ALTER TABLE prices ADD COLUMN foil_gap INTEGER NOT NULL DEFAULT 0")
    except sqlite3.OperationalError:
        pass  # column already exists
    return conn


# ── Import (Moxfield CSV) ─────────────────────────────────────────────────────

# Moxfield export headers we care about, matched case-insensitively.
_COLS = {
    "count": "count",
    "name": "name",
    "edition": "set_code",
    "condition": "condition",
    "language": "language",
    "foil": "foil",
    "collector number": "collector_number",
    "purchase price": "purchase_price",
}


def import_csv(text: str) -> dict:
    """Replaces the whole collection with the given Moxfield CSV export.
    Returns {imported, skipped, names}. Raises ValueError on an unusable file."""
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("empty file")
    field_map = {}
    for raw in reader.fieldnames:
        key = _COLS.get((raw or "").strip().strip('"').lower())
        if key:
            field_map[raw] = key
    mapped = set(field_map.values())
    if "name" not in mapped:
        raise ValueError("no 'Name' column — is this a Moxfield collection CSV export?")

    rows: dict[tuple, dict] = {}
    skipped = 0
    for raw_row in reader:
        row = {"count": 1, "name": "", "set_code": "", "collector_number": "",
               "condition": "", "language": "", "foil": "", "purchase_price": None}
        for raw_key, our_key in field_map.items():
            val = (raw_row.get(raw_key) or "").strip()
            if our_key == "count":
                try:
                    row["count"] = max(1, int(float(val)))
                except ValueError:
                    row["count"] = 1
            elif our_key == "purchase_price":
                try:
                    row["purchase_price"] = float(val.replace("$", "")) if val else None
                except ValueError:
                    row["purchase_price"] = None
            elif our_key == "foil":
                row["foil"] = val.lower()  # '', 'foil', 'etched'
            elif our_key == "set_code":
                row["set_code"] = val.upper()
            else:
                row[our_key] = val
        if not row["name"]:
            skipped += 1
            continue
        key = (row["name"], row["set_code"], row["collector_number"].lower(),
               row["foil"], row["condition"].lower(), row["language"].lower())
        if key in rows:
            rows[key]["count"] += row["count"]
        else:
            rows[key] = row

    if not rows:
        raise ValueError("no card rows found in the file")

    with _connect() as conn:
        conn.execute("DELETE FROM items")  # cascades to prices
        conn.executemany(
            "INSERT INTO items (name, set_code, collector_number, foil, condition, language, count, purchase_price) "
            "VALUES (:name, :set_code, :collector_number, :foil, :condition, :language, :count, :purchase_price)",
            list(rows.values()),
        )
    names = len({r["name"] for r in rows.values()})
    log.info("collection: imported %d items (%d unique names), %d rows skipped", len(rows), names, skipped)
    return {"imported": len(rows), "unique_names": names, "skipped": skipped}


def clear() -> None:
    with _connect() as conn:
        conn.execute("DELETE FROM items")


# ── Matcher ───────────────────────────────────────────────────────────────────

# Trait names -> lowercase text markers, from print_index. "promo" is too
# noisy as a *conflict* signal (stores sprinkle "promo" around), so conflicts
# only consider concrete physical-variant markers.
_CONFLICT_TRAITS = ("borderless", "showcase", "extended", "etched", "retro", "serialized")


def _listing_text(row: dict) -> str:
    return f"{row.get('name', '')} {row.get('extra_info', '')}".lower()


def _set_matches(text: str, identity: dict, set_code: str) -> bool:
    set_name = (identity.get("s") or "").lower()
    if set_name and set_name in text:
        # Substring collision guard: "Ixalan" also appears inside "The Lost
        # Caverns of Ixalan". Only accept the match if no LONGER known set
        # name containing ours is also in the text — the longer one is what
        # the listing actually names.
        for other in print_index.set_names():
            if other != set_name and set_name in other and other in text:
                break
        else:
            return True
    code = set_code.lower()
    if len(code) >= 3 and (f"[{code}]" in text or f"({code})" in text):
        return True
    return False


def _listing_traits(text: str) -> set[str]:
    """Concrete physical-variant markers a listing's text carries."""
    found = set()
    for trait in _CONFLICT_TRAITS:
        markers = print_index.TRAIT_MARKERS.get(trait, [trait])
        if any(m in text for m in markers):
            found.add(trait)
    return found


def _foil_matches(row: dict, item_foil: str) -> bool:
    wants_foil = item_foil in ("foil", "etched")
    return bool(row.get("foil")) == wants_foil


def match_item(item: dict, rows: list[dict]) -> dict:
    """Best listing for one collection item from that card name's /search rows.
    Returns {tier, price, store, url, listing, foil_gap} — tier 'none' when out
    of stock everywhere. foil_gap=True means the price is a proxy from the
    wrong finish (e.g. a foil item priced from a nonfoil listing because no
    foil is in stock anywhere)."""
    candidates = [r for r in rows
                  if not r.get("is_error") and float(r.get("price_val") or 0) > 0]
    if not candidates:
        return {"tier": "none", "price": None, "store": None, "url": None,
                "listing": None, "foil_gap": False}

    identity = print_index.lookup(item["set_code"], item["collector_number"])
    traits = set(identity.get("v") or []) if identity else set()
    if item["foil"] == "etched":
        traits.add("etched")
    # Only concrete physical markers participate in variant classification —
    # stores label prints with ONE of a print's traits ("Borderless"), not all.
    print_traits = traits & set(_CONFLICT_TRAITS)

    def cheapest(pool):
        return min(pool, key=lambda r: r["price_val"]) if pool else None

    best, tier = None, "any"
    if identity:
        # A card with exactly one printing in existence can't be mis-matched:
        # any listing IS that print, set text or not (foil still has to match).
        unique_print = print_index.printing_count(item["name"]) == 1
        # All printings in one set: a bare-name listing still proves the SET —
        # only the variant stays unknown, so such listings cap at "set" tier.
        single_set = print_index.set_codes_for(item["name"]) == {item["set_code"].upper()}
        exact_pool, set_pool = [], []
        for r in candidates:
            text = _listing_text(r)
            if not _foil_matches(r, item["foil"]):
                continue
            set_named = _set_matches(text, identity, item["set_code"])
            if not (set_named or unique_print or single_set):
                continue
            l_traits = _listing_traits(text)
            if not l_traits <= print_traits:
                continue  # listing is explicitly a DIFFERENT variant — worse than unconfirmed
            if unique_print:
                exact_pool.append(r)   # only one print exists — this is it
            elif print_traits and l_traits:
                exact_pool.append(r)   # special print, listing confirms (a subset of) its markers
            elif not print_traits and set_named:
                exact_pool.append(r)   # plain print, set named, no variant markers
            else:
                # set implied-or-named but the specific variant is unconfirmed
                set_pool.append(r)
        if exact_pool:
            best, tier = cheapest(exact_pool), "exact"
        elif set_pool:
            best, tier = cheapest(set_pool), "set"

    foil_gap = False
    if best is None:
        foil_pool = [r for r in candidates if _foil_matches(r, item["foil"])]
        best = cheapest(foil_pool) or cheapest(candidates)
        tier = "any"
        # Price taken from the wrong finish — a proxy, not a real quote.
        foil_gap = not _foil_matches(best, item["foil"])

    return {
        "tier": tier,
        "price": round(float(best["price_val"]), 2),
        "store": best.get("src", ""),
        "url": best.get("url", ""),
        "listing": best.get("name", ""),
        "foil_gap": foil_gap,
    }


# ── Pricer worker ─────────────────────────────────────────────────────────────

_worker_lock = threading.Lock()
_worker: threading.Thread | None = None
_pause_event = threading.Event()
_stop_event = threading.Event()
_progress = {"state": "idle", "done": 0, "total": 0, "current": "", "started": None, "error": ""}


def _pending_names(conn: sqlite3.Connection) -> list[str]:
    """Unique card names needing (re)pricing: never priced, or stale."""
    cutoff = datetime.now().timestamp() - REFRESH_AGE_DAYS * 86400
    cutoff_iso = datetime.fromtimestamp(cutoff).isoformat(timespec="seconds")
    cur = conn.execute(
        "SELECT DISTINCT i.name FROM items i "
        "LEFT JOIN prices p ON p.item_id = i.id "
        "WHERE p.item_id IS NULL OR p.checked_at < ? "
        "ORDER BY i.name",
        (cutoff_iso,),
    )
    return [r[0] for r in cur]


def _price_names(names: list[str]) -> None:
    """One batch: search, match every item of those names, commit."""
    resp = httpx.post(APP_SEARCH_URL, json={"buy_list": names}, timeout=SEARCH_TIMEOUT_S)
    resp.raise_for_status()
    results = resp.json().get("results", [])

    rows_by_name: dict[str, list[dict]] = {}
    for r in results:
        rows_by_name.setdefault(r.get("card", ""), []).append(r)

    now = datetime.now().isoformat(timespec="seconds")
    with _connect() as conn:
        for name in names:
            rows = rows_by_name.get(name, [])
            cur = conn.execute(
                "SELECT id, name, set_code, collector_number, foil FROM items WHERE name = ?",
                (name,),
            )
            for item_id, iname, set_code, number, foil in cur.fetchall():
                m = match_item(
                    {"name": iname, "set_code": set_code, "collector_number": number, "foil": foil},
                    rows,
                )
                conn.execute(
                    "INSERT OR REPLACE INTO prices (item_id, checked_at, tier, price, store, url, listing, foil_gap) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (item_id, now, m["tier"], m["price"], m["store"], m["url"], m["listing"],
                     int(m.get("foil_gap", False))),
                )


def _snapshot_value() -> None:
    with _connect() as conn:
        total, priced, items = conn.execute(
            "SELECT COALESCE(SUM(p.price * i.count), 0), "
            "       COUNT(p.item_id) FILTER (WHERE p.price IS NOT NULL), COUNT(i.id) "
            "FROM items i LEFT JOIN prices p ON p.item_id = i.id"
        ).fetchone()
        conn.execute(
            "INSERT OR REPLACE INTO value_snapshots (day, total_sgd, priced, items) VALUES (?, ?, ?, ?)",
            (date.today().isoformat(), round(total, 2), priced, items),
        )


def _worker_loop() -> None:
    try:
        with _connect() as conn:
            pending = _pending_names(conn)
        _progress.update(done=0, total=len(pending), error="")
        log.info("collection pricer: %d names to price", len(pending))

        for i in range(0, len(pending), BATCH_NAMES):
            if _stop_event.is_set():
                _progress["state"] = "stopped"
                return
            while _pause_event.is_set():
                _progress["state"] = "paused"
                if _stop_event.wait(1.0):
                    _progress["state"] = "stopped"
                    return
            _progress["state"] = "running"

            batch = pending[i:i + BATCH_NAMES]
            _progress["current"] = ", ".join(batch[:3]) + ("…" if len(batch) > 3 else "")
            try:
                _price_names(batch)
            except Exception as exc:  # noqa: BLE001 — one bad batch shouldn't kill the pass
                log.warning("collection pricer: batch failed (%s) — continuing: %s", batch[:2], exc)
                _progress["error"] = f"last batch error: {exc}"
            _progress["done"] = min(i + BATCH_NAMES, len(pending))

            if i + BATCH_NAMES < len(pending):
                if _stop_event.wait(BATCH_PAUSE_S):
                    _progress["state"] = "stopped"
                    return

        _snapshot_value()
        _progress.update(state="done", current="")
        log.info("collection pricer: pass complete (%d names)", len(pending))
    except Exception as exc:  # noqa: BLE001
        log.exception("collection pricer: worker crashed")
        _progress.update(state="error", error=str(exc))


def start_pricer() -> dict:
    global _worker
    with _worker_lock:
        if _worker is not None and _worker.is_alive():
            _pause_event.clear()  # treat start-while-paused as resume
            return status()
        _stop_event.clear()
        _pause_event.clear()
        _progress.update(state="running", done=0, total=0, current="",
                         started=datetime.now().isoformat(timespec="seconds"), error="")
        _worker = threading.Thread(target=_worker_loop, name="collection-pricer", daemon=True)
        _worker.start()
    return status()


def pause_pricer() -> dict:
    _pause_event.set()
    return status()


def stop_pricer() -> dict:
    _stop_event.set()
    _pause_event.clear()
    return status()


# ── Read API ──────────────────────────────────────────────────────────────────

def status() -> dict:
    with _connect() as conn:
        items, priced, total = conn.execute(
            "SELECT COUNT(i.id), COUNT(p.item_id) FILTER (WHERE p.price IS NOT NULL), "
            "       COALESCE(SUM(p.price * i.count), 0) "
            "FROM items i LEFT JOIN prices p ON p.item_id = i.id"
        ).fetchone()
        tiers = dict(conn.execute(
            "SELECT p.tier, COUNT(*) FROM prices p GROUP BY p.tier").fetchall())
        snapshots = conn.execute(
            "SELECT day, total_sgd FROM value_snapshots ORDER BY day DESC LIMIT 30").fetchall()
    return {
        "items": items,
        "priced": priced,
        "total_sgd": round(total, 2),
        "tiers": tiers,
        "pricer": dict(_progress),
        "index_ready": print_index.available(),
        "value_history": [[d, v] for d, v in reversed(snapshots)],
    }


def list_items(query: str = "", tier: str = "", offset: int = 0, limit: int = 100) -> dict:
    where, params = [], []
    if query:
        where.append("i.name LIKE ?")
        params.append(f"%{query}%")
    if tier == "unpriced":
        where.append("p.item_id IS NULL")
    elif tier:
        where.append("p.tier = ?")
        params.append(tier)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    with _connect() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) FROM items i LEFT JOIN prices p ON p.item_id = i.id {where_sql}",
            params,
        ).fetchone()[0]
        cur = conn.execute(
            "SELECT i.name, i.set_code, i.collector_number, i.foil, i.condition, i.count, "
            "       p.tier, p.price, p.store, p.url, p.listing, p.checked_at, p.foil_gap "
            f"FROM items i LEFT JOIN prices p ON p.item_id = i.id {where_sql} "
            "ORDER BY (p.price IS NULL), p.price * i.count DESC, i.name "
            "LIMIT ? OFFSET ?",
            [*params, limit, offset],
        )
        cols = ["name", "set_code", "collector_number", "foil", "condition", "count",
                "tier", "price", "store", "url", "listing", "checked_at", "foil_gap"]
        items = [dict(zip(cols, row)) for row in cur.fetchall()]

    # Exact-print Card Kingdom reference (USD) — foil-aware, straight from the
    # print index (pure dict lookups). Reference only, never part of totals.
    for it in items:
        it["ck_usd"], it["ck_url"] = None, None
        identity = print_index.lookup(it["set_code"], it["collector_number"])
        ck = (identity or {}).get("ck")
        if ck:
            if it["foil"] in ("foil", "etched"):
                it["ck_usd"], it["ck_url"] = ck.get("f"), ck.get("uf")
            else:
                it["ck_usd"], it["ck_url"] = ck.get("n"), ck.get("u")
    return {"total": total, "offset": offset, "items": items}


# ── xlsx export ───────────────────────────────────────────────────────────────

_TIER_LABELS = {"exact": "Exact print", "set": "Same set", "any": "Any print", "none": "No stock"}


def export_xlsx() -> bytes:
    """The whole collection as one xlsx sheet — same palette as export/excel.py."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    items = list_items(limit=1_000_000)["items"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"

    headers = ["Card", "Set", "No.", "Foil", "Condition", "Qty", "Match",
               "Unit (SGD)", "Total (SGD)", "CK (USD)", "Store", "Listing", "Checked"]
    header_font = Font(bold=True, color="6D28D9")
    header_fill = PatternFill("solid", fgColor="F3F4F6")
    header_border = Border(bottom=Side(style="thin", color="D1D5DB"))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.border = header_font, header_fill, header_border
    ws.freeze_panes = "A2"

    tier_color = {"exact": "15803D", "set": "B45309", "any": "4B5563", "none": "B91C1C"}
    grand_total = 0.0
    for r, it in enumerate(items, 2):
        total = (it["price"] or 0) * it["count"]
        grand_total += total
        match_label = _TIER_LABELS.get(it["tier"] or "", "Not priced")
        if it.get("foil_gap"):
            match_label += " — no foil in stock" if it["foil"] else " — only foil in stock"
        row = [
            it["name"], it["set_code"], it["collector_number"],
            it["foil"] or "", it["condition"], it["count"], match_label,
            it["price"], round(total, 2) if it["price"] is not None else None,
            it["ck_usd"], it["store"] or "", it["listing"] or "", it["checked_at"] or "",
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            if col == 1:
                cell.font = Font(bold=True, color="111827")
            elif col == 7:
                cell.font = Font(color=tier_color.get(it["tier"] or "", "6B7280"))
            elif col in (8, 9):
                cell.font = Font(color="065F46")
                cell.number_format = "0.00"
            elif col == 10:
                cell.font = Font(color="92400E")
                cell.number_format = "0.00"
        if it["url"]:
            link_cell = ws.cell(row=r, column=12)
            link_cell.hyperlink = it["url"]
            link_cell.font = Font(color="1D4ED8", underline="single")

    total_row = len(items) + 2
    label = ws.cell(row=total_row, column=8, value="Total:")
    label.font = Font(bold=True)
    label.alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=9, value=round(grand_total, 2))
    total_cell.font = Font(bold=True, color="065F46")
    total_cell.number_format = "0.00"

    widths = [34, 6, 7, 7, 11, 5, 24, 11, 11, 10, 20, 44, 19]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
