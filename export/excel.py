"""
export/excel.py — writes search result rows to an xlsx workbook.

Workbook layout
---------------
  Sheet 1 "Results"       — all top-5 results per card (unchanged from v1)
  Sheet 2 "Shopping Plan" — three shopping strategies (new in v2.1)

Results sheet schema:
    Card | Rank | Store | Listing Name | Set/Printing | Foil | Quality | Price (SGD)

Shopping Plan sheet layout:
    • Summary comparison table (strategies A / B / C side-by-side)
    • Strategy A block  — Cheapest Each (blue theme)
    • Strategy B block  — Best Shop greedy (amber theme)
    • Strategy C block  — Best Shop with price tolerance (green theme)
    Each block: store header rows → per-card rows (Card | Quality | Foil | Price | vs Cheapest)
    Not-found cards listed at the bottom of each block.

The Shopping Plan sheet is only written when write_excel() receives a
ShoppingPlan object.  If plan=None the workbook contains only the Results sheet,
preserving full backwards compatibility.
"""
from __future__ import annotations

import io
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from optimizer import ShoppingPlan, Strategy


# ═════════════════════════════════════════════════════════════════════════════
# Colour palette
# ═════════════════════════════════════════════════════════════════════════════

# Results sheet ───────────────────────────────────────────────────────────────
BG_HEADER = "F3F4F6"
FG_HEADER = "6D28D9"
FG_RANK1  = "B45309"
FG_RANK2  = "4B5563"
FG_RANK3  = "92400E"
FG_PRICE  = "065F46"
FG_STORE  = "3730A3"
FG_CARD   = "111827"
FG_INFO   = "6B7280"
FG_FOIL   = "6D28D9"
FG_ERROR  = "DC2626"
FG_LINK   = "1D4ED8"

_HEADER_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))

# Shopping Plan sheet — shared ────────────────────────────────────────────────
FG_CHEAPEST    = "15803D"   # dark green — this is the cheapest price
FG_PREMIUM     = "B45309"   # amber — above cheapest
FG_NOT_FOUND   = "B91C1C"   # red text
BG_NOT_FOUND   = "FEF2F2"   # very light red fill
FG_WHITE       = "FFFFFF"
FG_DARK        = "111827"
BG_COL_HEADER  = "E5E7EB"   # light grey for column header rows in plan sheet
FG_COL_HEADER  = "374151"

# Summary table
BG_SUMMARY_HDR = "374151"   # dark grey
FG_SUMMARY_HDR = "F9FAFB"
BG_SUMMARY_A   = "EFF6FF"   # light blue tint
BG_SUMMARY_B   = "FFFBEB"   # light amber tint
BG_SUMMARY_C   = "F0FDF4"   # light green tint

# Strategy A — blue theme
BG_STRAT_A       = "1E3A5F"   # dark navy header
BG_STRAT_A_DESC  = "EFF6FF"   # very light blue description row
BG_STRAT_A_STORE = "BFDBFE"   # medium light blue store header
FG_STRAT_A_STORE = "1E3A5F"   # dark navy text on store header

# Strategy B — amber theme
BG_STRAT_B       = "78350F"   # dark amber header
BG_STRAT_B_DESC  = "FFFBEB"   # very light amber description row
BG_STRAT_B_STORE = "FDE68A"   # medium light amber store header
FG_STRAT_B_STORE = "78350F"   # dark amber text on store header

# Strategy C — green theme
BG_STRAT_C       = "14532D"   # dark green header
BG_STRAT_C_DESC  = "F0FDF4"   # very light green description row
BG_STRAT_C_STORE = "BBF7D0"   # medium light green store header
FG_STRAT_C_STORE = "14532D"   # dark green text on store header


# ═════════════════════════════════════════════════════════════════════════════
# Results sheet (unchanged logic from v1)
# ═════════════════════════════════════════════════════════════════════════════

RESULT_HEADERS    = ["Card", "Rank", "Store", "Listing Name", "Set / Printing", "Foil", "Quality", "Price (SGD)"]
RESULT_COL_WIDTHS = [26, 7, 24, 44, 32, 7, 10, 14]


def _write_results_sheet(ws, rows: list[dict]) -> None:
    ws.title = "Results"

    hdr_fill = PatternFill("solid", fgColor=BG_HEADER)
    hdr_font = Font(bold=True, color=FG_HEADER, name="Calibri")
    for col_idx, (header, width) in enumerate(zip(RESULT_HEADERS, RESULT_COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for row_num, row in enumerate(rows, 2):
        is_error = row.get("is_error", False)
        rank_n   = row.get("rank_n", 0)
        rank     = row.get("rank", "")
        url      = row.get("url", "")

        def _cell(col: int, value, fg: str, bold: bool = False, underline: str = "none"):
            c = ws.cell(row=row_num, column=col, value=value)
            c.font = Font(color=fg, bold=bold, name="Calibri", underline=underline)
            c.alignment = Alignment(horizontal="left", vertical="center")
            return c

        _cell(1, row["card"], FG_CARD, bold=True)

        if is_error:
            rank_fg = FG_ERROR
        elif rank_n == 1:
            rank_fg = FG_RANK1
        elif rank_n == 2:
            rank_fg = FG_RANK2
        elif rank_n == 3:
            rank_fg = FG_RANK3
        else:
            rank_fg = FG_INFO
        _cell(2, rank, rank_fg, bold=rank_n in (1, 2, 3))

        _cell(3, row["src"], FG_STORE)

        name_cell = _cell(4, row["name"], FG_LINK if url and not is_error else FG_CARD)
        if url and not is_error:
            name_cell.hyperlink = url
            name_cell.font = Font(color=FG_LINK, underline="single", name="Calibri")

        _cell(5, row["extra_info"], FG_INFO)

        foil_val = "✦" if row.get("foil") else ""
        _cell(6, foil_val, FG_FOIL, bold=bool(row.get("foil")))

        _cell(7, row["quality"], FG_INFO)

        if is_error:
            _cell(8, row["price"], FG_ERROR)
        else:
            _cell(8, row["price_val"], FG_PRICE, bold=True)
            ws.cell(row=row_num, column=8).number_format = '"SGD "#,##0.00'

        ws.row_dimensions[row_num].height = 16


# ═════════════════════════════════════════════════════════════════════════════
# Shopping Plan sheet
# ═════════════════════════════════════════════════════════════════════════════

# Plan sheet uses 5 columns (A–E).
PLAN_COL_WIDTHS = [30, 10, 6, 14, 24]   # Card | Quality | Foil | Price | vs Cheapest
NUM_PLAN_COLS   = 5


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(color: str, bold: bool = False, size: int = 10, italic: bool = False) -> Font:
    return Font(color=color, bold=bold, size=size, name="Calibri", italic=italic)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _merge_row(ws, row: int, value: str, bg: str, fg: str,
               bold: bool = True, size: int = 10, height: int = 18,
               italic: bool = False) -> None:
    """Write a full-width merged label row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_PLAN_COLS)
    cell = ws.cell(row=row, column=1, value=value)
    cell.fill    = _fill(bg)
    cell.font    = _font(fg, bold=bold, size=size, italic=italic)
    cell.alignment = _align(h="left", v="center")
    ws.row_dimensions[row].height = height


def _blank_row(ws, row: int, height: int = 6) -> None:
    ws.row_dimensions[row].height = height


def _write_summary_table(ws, plan: "ShoppingPlan", start_row: int) -> int:
    """Write the 3-row comparison table. Returns next available row."""
    row = start_row

    # Title
    _merge_row(ws, row, "🛒  GISHATH SHOPPING PLAN", BG_STRAT_A, FG_WHITE,
               bold=True, size=13, height=26)
    row += 1
    _blank_row(ws, row)
    row += 1

    # Column headers
    headers = ["Strategy", "Total (SGD)", "Stores", "Cards Found", "Premium vs Min"]
    for col, header in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=header)
        c.fill      = _fill(BG_SUMMARY_HDR)
        c.font      = _font(FG_SUMMARY_HDR, bold=True)
        c.alignment = _align(h="center" if col > 1 else "left")
    ws.row_dimensions[row].height = 18
    row += 1

    strategies = [plan.strategy_a, plan.strategy_b, plan.strategy_c]
    tints       = [BG_SUMMARY_A, BG_SUMMARY_B, BG_SUMMARY_C]
    min_cost    = plan.min_possible_cost

    for strat, tint in zip(strategies, tints):
        found_str = f"{strat.cards_found} / {plan.total_cards_searched}"

        if strat.total_cost == 0 or min_cost == 0:
            premium_str = "—"
        elif strat is plan.strategy_a:
            premium_str = "— (baseline)"
        else:
            diff = strat.total_cost - min_cost
            pct  = (diff / min_cost * 100) if min_cost else 0
            if diff <= 0.005:
                premium_str = "same"
            else:
                premium_str = f"+{diff:.2f}  (+{pct:.1f}%)"

        values = [
            strat.label,
            strat.total_cost,
            strat.store_count,
            found_str,
            premium_str,
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(tint)
            c.alignment = _align(h="center" if col > 1 else "left")
            c.font      = _font(FG_DARK, bold=(col == 1))
            if col == 2 and isinstance(val, float):
                c.number_format = '"SGD "#,##0.00'
            # Premium column: colour the text
            if col == 5 and isinstance(val, str):
                if val.startswith("+"):
                    c.font = _font(FG_PREMIUM, bold=False)
                elif val in ("same", "— (baseline)"):
                    c.font = _font(FG_CHEAPEST, bold=False)

        ws.row_dimensions[row].height = 18
        row += 1

    return row


def _write_strategy_block(
    ws,
    strategy: "Strategy",
    plan: "ShoppingPlan",
    start_row: int,
    bg_header: str,
    bg_desc: str,
    bg_store: str,
    fg_store: str,
) -> int:
    """Write one full strategy block. Returns next available row."""
    row = start_row

    _blank_row(ws, row, height=10)
    row += 1

    # Strategy heading
    _merge_row(ws, row, f"  {strategy.label}", bg_header, FG_WHITE,
               bold=True, size=11, height=22)
    row += 1

    # Description
    _merge_row(ws, row, f"  {strategy.description}", bg_desc, FG_DARK,
               bold=False, size=9, height=14, italic=True)
    row += 1

    # If no cards found
    if not strategy.groups:
        _merge_row(ws, row, "  No results to display.", bg_desc, FG_INFO,
                   bold=False, size=9, height=14, italic=True)
        row += 1
        if strategy.not_found:
            row = _write_not_found(ws, strategy.not_found, row, bg_store)
        return row

    # Per-store sections
    for group in strategy.groups:
        _blank_row(ws, row, height=4)
        row += 1

        # Store header row
        store_label = (
            f"  {group.store}   ·   {len(group.assignments)} card"
            f"{'s' if len(group.assignments) != 1 else ''}   ·   SGD {group.total:,.2f}"
        )
        _merge_row(ws, row, store_label, bg_store, fg_store,
                   bold=True, size=10, height=18)
        row += 1

        # Column headers for this section
        col_hdrs = ["  Card", "Quality", "Foil", "Price (SGD)", "vs Cheapest"]
        for col, hdr in enumerate(col_hdrs, 1):
            c = ws.cell(row=row, column=col, value=hdr)
            c.fill      = _fill(BG_COL_HEADER)
            c.font      = _font(FG_COL_HEADER, bold=True, size=9)
            c.alignment = _align(h="center" if col > 1 else "left")
        ws.row_dimensions[row].height = 15
        row += 1

        # Card rows
        for assignment in group.assignments:
            # Card name (indent for readability)
            c = ws.cell(row=row, column=1, value=f"  {assignment.card}")
            c.font      = _font(FG_DARK, bold=False)
            c.alignment = _align()

            # Quality
            c = ws.cell(row=row, column=2, value=assignment.quality)
            c.font      = _font(FG_INFO)
            c.alignment = _align(h="center")

            # Foil
            foil_str = "✦" if assignment.foil else ""
            c = ws.cell(row=row, column=3, value=foil_str)
            c.font      = _font(FG_FOIL if assignment.foil else FG_INFO, bold=assignment.foil)
            c.alignment = _align(h="center")

            # Price — hyperlinked if URL available
            price_cell = ws.cell(row=row, column=4, value=assignment.price)
            price_cell.number_format = '"SGD "#,##0.00'
            price_cell.alignment     = _align(h="right")
            if assignment.url:
                price_cell.hyperlink = assignment.url
                price_cell.font = Font(
                    color=FG_LINK, bold=True, underline="single",
                    size=10, name="Calibri"
                )
            else:
                price_cell.font = _font(FG_CHEAPEST if assignment.is_cheapest else FG_PREMIUM, bold=True)

            # vs Cheapest
            if assignment.is_cheapest:
                vs_val  = "cheapest  ✓"
                vs_font = _font(FG_CHEAPEST)
            else:
                diff    = assignment.premium
                pct     = (diff / assignment.cheapest_price * 100) if assignment.cheapest_price else 0
                vs_val  = f"+{diff:.2f}  (+{pct:.0f}%)"
                vs_font = _font(FG_PREMIUM)

            c = ws.cell(row=row, column=5, value=vs_val)
            c.font      = vs_font
            c.alignment = _align(h="left")

            ws.row_dimensions[row].height = 15
            row += 1

    # Not-found section
    if strategy.not_found:
        _blank_row(ws, row, height=4)
        row += 1
        row = _write_not_found(ws, strategy.not_found, row, bg_store)

    return row


def _write_not_found(ws, not_found: list[str], start_row: int, bg_hint: str) -> int:
    """Write a compact 'Not Found' section. Returns next available row."""
    row = start_row

    label = f"  ✗  NOT FOUND — {len(not_found)} card{'s' if len(not_found) != 1 else ''} could not be sourced"
    _merge_row(ws, row, label, BG_NOT_FOUND, FG_NOT_FOUND, bold=True, size=9, height=16)
    row += 1

    for card_name in not_found:
        c = ws.cell(row=row, column=1, value=f"    {card_name}")
        c.font      = _font(FG_NOT_FOUND, italic=True, size=9)
        c.alignment = _align()
        ws.row_dimensions[row].height = 13
        row += 1

    return row


def _write_shopping_plan_sheet(ws, plan: "ShoppingPlan") -> None:
    """Populate a worksheet with the full Shopping Plan."""
    ws.title = "Shopping Plan"

    # Set column widths
    for col_idx, width in enumerate(PLAN_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Summary table
    current_row = _write_summary_table(ws, plan, start_row=1)

    # Strategy blocks
    strategy_themes = [
        (plan.strategy_a, BG_STRAT_A, BG_STRAT_A_DESC, BG_STRAT_A_STORE, FG_STRAT_A_STORE),
        (plan.strategy_b, BG_STRAT_B, BG_STRAT_B_DESC, BG_STRAT_B_STORE, FG_STRAT_B_STORE),
        (plan.strategy_c, BG_STRAT_C, BG_STRAT_C_DESC, BG_STRAT_C_STORE, FG_STRAT_C_STORE),
    ]
    for strat, bg_h, bg_d, bg_s, fg_s in strategy_themes:
        current_row = _write_strategy_block(
            ws, strat, plan, current_row,
            bg_header=bg_h, bg_desc=bg_d, bg_store=bg_s, fg_store=fg_s,
        )

    # Freeze the title row
    ws.freeze_panes = "A2"


# ═════════════════════════════════════════════════════════════════════════════
# Public entry point
# ═════════════════════════════════════════════════════════════════════════════

def write_excel(rows: list[dict], plan=None) -> bytes:
    """Build an xlsx workbook and return as raw bytes.

    Args:
        rows: Display rows from format_results() — top-5 per card, no hidden rows.
        plan: Optional ShoppingPlan from compute_plan().  When provided, a second
              'Shopping Plan' sheet is added to the workbook.

    Returns:
        Raw xlsx bytes suitable for send_file() or writing to disk.
    """
    wb = Workbook()

    # Sheet 1 — Results
    _write_results_sheet(wb.active, rows)

    # Sheet 2 — Shopping Plan (only if plan was computed)
    if plan is not None:
        _write_shopping_plan_sheet(wb.create_sheet(), plan)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
