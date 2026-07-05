"""Regression test for the bug Trevor caught live (2026-07-01): optimize's own
fact-check correctly determined the Zinnia, Valley's Voice deck's original
"archetype"/"rationale" label ("Izzet dice-rolling chaos spellslinger" —
written at the select stage, before the deck existed or was fact-checked)
didn't match the commander's real printed text, and built the actual deck
around her real offspring/power-1 mechanic instead — but the WRONG original
label still shipped verbatim in the email and xlsx, because emailer.py/
export.py read deck.concept.archetype/rationale (the select-stage's
unverified first guess) directly, never deck's fact-checked final version.

Fixed by having optimize.md/OPTIMIZE_JSON_SCHEMA return final_archetype/
final_summary (corrected or confirmed-accurate) and having export.py/
emailer.py display those instead. This test builds a DeckResult with a
deliberately WRONG concept.archetype/rationale and a deliberately DIFFERENT
final_archetype/final_summary (as if optimize had corrected it), then checks
the wrong text never appears anywhere in the plain email body, HTML email
body, or the xlsx Gameplan sheet — only the corrected text does.

Run manually: cd gishath-local-v2 && python3 -m deck_engine.tests.test_final_archetype_display
"""
from __future__ import annotations

import io
import sys

from openpyxl import load_workbook

from .. import emailer, export
from ..agent_pipeline import DeckResult
from ..concept_selector import ConceptChoice
from ..pricing import PricingOutcome
from ..scryfall_cache import ValidationResult

WRONG_ARCHETYPE = "Izzet dice-rolling chaos spellslinger"
WRONG_RATIONALE = "Zinnia turns every roll into extra copies of instants/sorceries and cascading die-roll payoffs"
CORRECTED_ARCHETYPE = "Izzet offspring ETB-doubling aristocrats"
CORRECTED_SUMMARY = "Zinnia grows off power-1 offspring tokens while doubled ETBs and sac-outlet pings drain the table"


def _build_fake_deck() -> DeckResult:
    concept = ConceptChoice(
        commander="Zinnia, Valley's Voice", archetype=WRONG_ARCHETYPE, rationale=WRONG_RATIONALE,
        color_identity=["U", "R"], oracle_text="(real oracle text — offspring/power-1, not dice)",
    )
    validation = ValidationResult(commander=concept.commander, card_count=100)
    return DeckResult(
        concept=concept, cards=["Sol Ring"] * 99, validation=validation,
        changes_made="No changes.", early_game="early", mid_game="mid", late_game="late",
        final_archetype=CORRECTED_ARCHETYPE, final_summary=CORRECTED_SUMMARY,
    )


def _assert_no_leak(label: str, text: str) -> list[str]:
    problems = []
    if WRONG_ARCHETYPE in text:
        problems.append(f"{label}: still contains the WRONG archetype label")
    if WRONG_RATIONALE in text:
        problems.append(f"{label}: still contains the WRONG rationale text")
    if CORRECTED_ARCHETYPE not in text:
        problems.append(f"{label}: missing the corrected archetype label")
    if CORRECTED_SUMMARY not in text:
        problems.append(f"{label}: missing the corrected summary")
    return problems


def main() -> int:
    deck = _build_fake_deck()
    spend_summary = {"total_cost_usd": 1.0, "total_turns": 10, "total_duration_ms": 1000}
    pricing = PricingOutcome(plan=None, available=False, error="test: pricing skipped")
    price_summary = {"total": 0.0, "priced_count": 0, "unpriced_count": 100, "top_expensive": []}
    last_decks: list[str] = []

    problems: list[str] = []
    problems += _assert_no_leak("plain email body", emailer._plain_success_body(  # noqa: SLF001
        deck, spend_summary, "PASSED", price_summary, last_decks, include_diagnostics=True))
    problems += _assert_no_leak("HTML email body", emailer._html_success_body(  # noqa: SLF001
        deck, spend_summary, "PASSED", True, price_summary, last_decks, None, include_diagnostics=True))

    xlsx_bytes = export.write_deck_excel(deck, pricing, cache={})
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    gameplan_text = "\n".join(
        str(cell.value) for row in wb["Gameplan"].iter_rows() for cell in row if cell.value is not None
    )
    problems += _assert_no_leak("xlsx Gameplan sheet", gameplan_text)

    if problems:
        print(f"FAILED: {problems}", file=sys.stderr)
        return 1

    print("OK: corrected final_archetype/final_summary ship everywhere; the original wrong label doesn't.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
