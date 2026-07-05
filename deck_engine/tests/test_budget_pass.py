"""Tests for the §3.4 budget pass (PRD v4 amendment, added 2026-07-03 after the
Arcades run shipped an SGD 1,023.86 deck).

Covers: the happy swap path (over-cap card swapped, re-priced, tagged), the
ship-flagged path (no usable substitute after the attempt budget), the
never-touch-compliant-cards vetting, the commander flag path, the select-stage
commander price re-pick, the post-swap synergy re-check note, and the
pricing-unavailable skip. All model calls and /search pricing are mocked —
same conventions as test_v4_amendment.py.

Run manually: cd gishath-local-v2 && python3 -m deck_engine.tests.test_budget_pass
"""
from __future__ import annotations

import sys
from types import SimpleNamespace
from unittest import mock

from .. import budget_pass, config
from ..agent_pipeline import DeckResult
from ..concept_selector import ConceptChoice
from ..pricing import PricingOutcome, cheapest_by_card, deck_price_summary
from ..scryfall_cache import ValidationResult

_CAP = config.MAX_CARD_PRICE_SGD  # 150 by default; tests use it symbolically, not hardcoded


def _mk_pricing(prices: dict[str, tuple[float, str]]) -> PricingOutcome:
    """PricingOutcome whose plan duck-types exactly what cheapest_by_card() reads."""
    assignments = [SimpleNamespace(card=name, price=price, store=store)
                   for name, (price, store) in prices.items()]
    plan = SimpleNamespace(strategy_a=SimpleNamespace(all_assignments=assignments))
    return PricingOutcome(plan=plan, available=True)


def _mk_deck(cards: list[str], commander: str = "Test Commander") -> DeckResult:
    concept = ConceptChoice(
        commander=commander, archetype="Test archetype", rationale="r",
        color_identity=["G"], oracle_text="Whenever a Wall enters, draw a card.",
        mechanic_tokens=["wall", "defender"],
    )
    validation = ValidationResult(commander=commander, card_count=len(cards) + 1)
    deck = DeckResult(
        concept=concept, cards=list(cards), validation=validation,
        changes_made="", early_game="", mid_game="", late_game="",
        final_archetype="Test archetype", final_summary="test summary",
    )
    for c in cards:
        deck.card_tags[c.strip().lower()] = {"role": "Synergy piece", "phase": "mid"}
    return deck


class _FakeClaudeResult:
    def __init__(self, payload: dict):
        self._payload = payload

    def parsed_json(self) -> dict:
        return self._payload


def _passthrough_validate(run_id, concept, cards, cache, stage_prefix, max_attempts):
    """Stand-in for _validate_and_repair: everything the swap produced is 'valid'."""
    return cards, ValidationResult(commander=concept.commander, card_count=len(cards) + 1)


# Synergy-friendly fake cache: every card matches the "wall" token unless named otherwise.
_CACHE = {
    "expensive staple": {"type_line": "Creature — Wall", "oracle_text": "Defender. Walls you control get +0/+3."},
    "budget wall": {"type_line": "Creature — Wall", "oracle_text": "Defender."},
    "generic value card": {"type_line": "Sorcery", "oracle_text": "Draw three cards."},
    "cheap card": {"type_line": "Creature — Wall", "oracle_text": "Defender, reach."},
    "test commander": {"type_line": "Legendary Creature — Dragon", "oracle_text": "Walls matter."},
}


def test_happy_swap_path() -> list[str]:
    problems = []
    deck = _mk_deck(["Expensive Staple", "Cheap Card"])
    pricing = _mk_pricing({
        "Test Commander": (2.0, "Store A"),
        "Expensive Staple": (_CAP + 250.0, "Store A"),
        "Cheap Card": (1.0, "Store B"),
    })
    swap_payload = {"swaps": [{"remove": "Expensive Staple", "add": "Budget Wall",
                               "reason": "same role, in print", "role": "Synergy piece", "phase": "mid"}]}

    def fake_fetch(names):
        return _mk_pricing({n: (3.5, "Store C") for n in names})

    with mock.patch.object(budget_pass.claude_cli, "run", return_value=_FakeClaudeResult(swap_payload)), \
         mock.patch.object(budget_pass, "_validate_and_repair", _passthrough_validate), \
         mock.patch.object(budget_pass.pricing_mod, "fetch_prices", fake_fetch):
        outcome = budget_pass.enforce_card_cap("test-run", deck, pricing, _CACHE)

    if not outcome.ran:
        problems.append("expected outcome.ran=True with pricing available")
    if "Budget Wall" not in deck.cards or "Expensive Staple" in deck.cards:
        problems.append(f"expected the over-cap card swapped out in place, got {deck.cards}")
    if len(outcome.swaps_made) != 1:
        problems.append(f"expected exactly 1 recorded swap, got {outcome.swaps_made}")
    if outcome.over_budget:
        problems.append(f"expected nothing left over cap, got {outcome.over_budget}")
    # Re-price overlay must flow through to every downstream consumer:
    cheapest = cheapest_by_card(pricing)
    if cheapest.get("budget wall") != (3.5, "Store C"):
        problems.append(f"expected the swapped-in card's fresh price in cheapest_by_card, got {cheapest.get('budget wall')}")
    summary = deck_price_summary(pricing, [deck.concept.commander] + deck.cards)
    if abs(summary["total"] - (2.0 + 3.5 + 1.0)) > 0.01:
        problems.append(f"expected the deck total to reflect the swap (6.50), got {summary['total']}")
    # Tags must follow the swap so the breakdown sheet stays complete:
    if deck.card_tags.get("budget wall", {}).get("role") != "Synergy piece":
        problems.append("expected the swapped-in card to inherit a role tag from the swap response")
    if "expensive staple" in deck.card_tags:
        problems.append("expected the removed card's tag to be dropped")
    return problems


def test_unfixable_ships_flagged_never_raises() -> list[str]:
    """Trevor's explicit call: never fail a run over a pricing rule. The model keeps
    proposing a substitute that ALSO prices over cap — after the attempt budget the
    deck ships with the breach flagged."""
    problems = []
    deck = _mk_deck(["Expensive Staple", "Cheap Card"])
    pricing = _mk_pricing({
        "Test Commander": (2.0, "Store A"),
        "Expensive Staple": (_CAP + 250.0, "Store A"),
        "Cheap Card": (1.0, "Store B"),
    })
    call_count = {"n": 0}

    def fake_run(*args, **kwargs):
        call_count["n"] += 1
        return _FakeClaudeResult({"swaps": [{
            "remove": "Expensive Staple" if call_count["n"] == 1 else "Also Expensive",
            "add": "Also Expensive" if call_count["n"] == 1 else "Still Expensive",
            "reason": "r", "role": "Synergy piece", "phase": "mid"}]})

    def fake_fetch(names):  # every substitute also prices over cap
        return _mk_pricing({n: (_CAP + 100.0, "Store C") for n in names})

    with mock.patch.object(budget_pass.claude_cli, "run", fake_run), \
         mock.patch.object(budget_pass, "_validate_and_repair", _passthrough_validate), \
         mock.patch.object(budget_pass.pricing_mod, "fetch_prices", fake_fetch):
        try:
            outcome = budget_pass.enforce_card_cap("test-run", deck, pricing, _CACHE)
        except Exception as exc:  # noqa: BLE001
            return [f"budget pass raised ({exc!r}) — it must NEVER fail a run for budget reasons"]

    if call_count["n"] != config.MAX_BUDGET_REPAIR_ATTEMPTS:
        problems.append(f"expected exactly {config.MAX_BUDGET_REPAIR_ATTEMPTS} swap attempts, got {call_count['n']}")
    if not outcome.over_budget:
        problems.append("expected the still-over-cap card on the over_budget flag list")
    if len(deck.cards) != 2:
        problems.append(f"deck size changed across failed swap attempts: {deck.cards}")
    return problems


def test_model_cannot_touch_compliant_cards() -> list[str]:
    """Same trust-nothing posture as T3: a swap targeting a card that is NOT a
    violation must be ignored, whatever the model returns."""
    problems = []
    deck = _mk_deck(["Expensive Staple", "Cheap Card"])
    pricing = _mk_pricing({
        "Test Commander": (2.0, "Store A"),
        "Expensive Staple": (_CAP + 50.0, "Store A"),
        "Cheap Card": (1.0, "Store B"),
    })
    rogue_payload = {"swaps": [
        {"remove": "Cheap Card", "add": "Sneaky Replacement", "reason": "rogue",
         "role": "Ramp", "phase": "early"},  # NOT a violation — must be vetted out
        {"remove": "Expensive Staple", "add": "Budget Wall", "reason": "legit",
         "role": "Synergy piece", "phase": "mid"},
    ]}

    def fake_fetch(names):
        return _mk_pricing({n: (3.5, "Store C") for n in names})

    with mock.patch.object(budget_pass.claude_cli, "run", return_value=_FakeClaudeResult(rogue_payload)), \
         mock.patch.object(budget_pass, "_validate_and_repair", _passthrough_validate), \
         mock.patch.object(budget_pass.pricing_mod, "fetch_prices", fake_fetch):
        outcome = budget_pass.enforce_card_cap("test-run", deck, pricing, _CACHE)

    if "Cheap Card" not in deck.cards:
        problems.append("a compliant card was removed — the vetting failed")
    if "Sneaky Replacement" in deck.cards:
        problems.append("a rogue (non-violation) swap's add made it into the deck")
    if "Budget Wall" not in deck.cards:
        problems.append("the legitimate swap should still have been applied")
    if not any("non-violating" in n for n in outcome.notes):
        problems.append(f"expected a note about the ignored rogue swap, got {outcome.notes}")
    return problems


def test_over_cap_commander_is_flagged_not_swapped() -> list[str]:
    problems = []
    deck = _mk_deck(["Cheap Card"])
    pricing = _mk_pricing({
        "Test Commander": (_CAP + 150.0, "Store A"),
        "Cheap Card": (1.0, "Store B"),
    })
    with mock.patch.object(budget_pass.claude_cli, "run") as run_mock:
        outcome = budget_pass.enforce_card_cap("test-run", deck, pricing, _CACHE)

    if run_mock.called:
        problems.append("no 99 cards were over cap — no swap model call should have been made")
    if not any(c == "Test Commander" for c, _ in outcome.over_budget):
        problems.append(f"expected the commander on the over_budget list, got {outcome.over_budget}")
    if deck.concept.commander != "Test Commander":
        problems.append("the commander must never be swapped by the budget pass")
    return problems


def test_pricing_unavailable_skips_cleanly() -> list[str]:
    problems = []
    deck = _mk_deck(["Expensive Staple"])
    pricing = PricingOutcome(plan=None, available=False, error="app down")
    with mock.patch.object(budget_pass.claude_cli, "run") as run_mock:
        outcome = budget_pass.enforce_card_cap("test-run", deck, pricing, _CACHE)
    if outcome.ran:
        problems.append("expected ran=False when pricing is unavailable")
    if run_mock.called:
        problems.append("no model call should happen when the cap can't be evaluated")
    if outcome.over_budget or outcome.swaps_made:
        problems.append("nothing should be flagged or swapped without prices")
    return problems


def test_synergy_recheck_notes_a_dip() -> list[str]:
    """A budget swap that replaces a synergy piece with generic filler must produce
    a synergy_note (check-only — the PRD is explicit there's no repair loop here)."""
    problems = []
    deck = _mk_deck(["Expensive Staple"])
    pricing = _mk_pricing({
        "Test Commander": (2.0, "Store A"),
        "Expensive Staple": (_CAP + 50.0, "Store A"),
    })
    swap_payload = {"swaps": [{"remove": "Expensive Staple", "add": "Generic Value Card",
                               "reason": "cheap", "role": "Card draw", "phase": "mid"}]}

    def fake_fetch(names):
        return _mk_pricing({n: (2.0, "Store C") for n in names})

    # Threshold 1 with a 1-card deck whose only nonland card matches nothing -> dip.
    with mock.patch.object(budget_pass.claude_cli, "run", return_value=_FakeClaudeResult(swap_payload)), \
         mock.patch.object(budget_pass, "_validate_and_repair", _passthrough_validate), \
         mock.patch.object(budget_pass.pricing_mod, "fetch_prices", fake_fetch), \
         mock.patch.object(config, "SYNERGY_GATE_THRESHOLD", 1):
        outcome = budget_pass.enforce_card_cap("test-run", deck, pricing, _CACHE)

    if not outcome.synergy_note:
        problems.append("expected a synergy_note after swapping the only on-mechanic card for generic filler")
    if "Generic Value Card" not in deck.cards:
        problems.append("the swap itself should still stand — the re-check reports, never reverts")
    return problems


def test_select_stage_commander_price_repick() -> list[str]:
    """concept_selector must re-pick when price_check says the commander is over cap,
    and accept a later within-cap pick."""
    from .. import concept_selector

    problems = []
    cache = {
        "pricey dragon": {"name": "Pricey Dragon", "type_line": "Legendary Creature — Dragon",
                          "oracle_text": "Flying.", "color_identity": ["R"]},
        "budget knight": {"name": "Budget Knight", "type_line": "Legendary Creature — Knight",
                          "oracle_text": "First strike.", "color_identity": ["W"]},
    }
    picks = iter([
        {"commander": "Pricey Dragon", "archetype": "dragons", "rationale": "r"},
        {"commander": "Budget Knight", "archetype": "knights", "rationale": "r"},
    ])
    prices = {"Pricey Dragon": _CAP + 500.0, "Budget Knight": 3.0}

    def fake_run(*args, **kwargs):
        return _FakeClaudeResult(next(picks))

    with mock.patch.object(concept_selector.claude_cli, "run", fake_run), \
         mock.patch.object(concept_selector.run_log, "recent_commanders", return_value=[]), \
         mock.patch.object(concept_selector.run_log, "recent_archetypes", return_value=set()), \
         mock.patch.object(concept_selector.synergy_check, "extract_mechanic_tokens", return_value=["knight"]):
        choice = concept_selector.select_concept(
            "test-run", cache=cache, max_attempts=3, price_check=lambda name: prices.get(name),
        )

    if choice.commander != "Budget Knight":
        problems.append(f"expected the over-cap commander re-picked, got {choice.commander}")
    return problems


def test_email_and_xlsx_budget_callouts() -> list[str]:
    """The budget block must appear in both email bodies (friends see it too — it's
    deck content, not diagnostics) and on the xlsx Gameplan sheet; and must be
    entirely absent when the pass did nothing."""
    import io

    from openpyxl import load_workbook

    from .. import emailer, export

    problems = []
    deck = _mk_deck(["Budget Wall"])
    outcome = budget_pass.BudgetOutcome(
        ran=True, cap=_CAP,
        swaps_made=[("Expensive Staple", _CAP + 250.0, "Budget Wall", 3.5, "same role, in print")],
        over_budget=[("Stubborn Card", _CAP + 20.0)],
    )
    spend = {"total_cost_usd": 1.0, "total_turns": 5, "total_duration_ms": 1000, "tools_used": []}
    price_summary = {"total": 10.0, "priced_count": 2, "unpriced_count": 0, "top_expensive": [("Budget Wall", 3.5)]}

    plain_clean = emailer._plain_success_body(  # noqa: SLF001
        deck, spend, "PASSED", price_summary, [], include_diagnostics=False, budget=outcome)
    html_clean = emailer._html_success_body(  # noqa: SLF001
        deck, spend, "PASSED", True, price_summary, [], None, include_diagnostics=False, budget=outcome)
    for body, label in ((plain_clean, "plain"), (html_clean, "html")):
        if "Budget Wall" not in body or "Expensive Staple" not in body:
            problems.append(f"{label} body missing the budget swap callout (friends' copy included)")
        if "Stubborn Card" not in body:
            problems.append(f"{label} body missing the over-budget flag")

    plain_no_budget = emailer._plain_success_body(  # noqa: SLF001
        deck, spend, "PASSED", price_summary, [], include_diagnostics=False,
        budget=budget_pass.BudgetOutcome(ran=True, cap=_CAP))
    if "Budget pass" in plain_no_budget or "OVER BUDGET" in plain_no_budget:
        problems.append("budget block should be absent when the pass made no swaps and flagged nothing")

    pricing = PricingOutcome(plan=None, available=False, error="test")
    wb = load_workbook(io.BytesIO(export.write_deck_excel(deck, pricing, cache={}, budget=outcome)))
    gameplan_labels = [r[0] for r in wb["Gameplan"].iter_rows(values_only=True)]
    if not any(label and "Budget pass" in str(label) for label in gameplan_labels):
        problems.append(f"Gameplan sheet missing the budget-pass row, got labels: {gameplan_labels}")
    if not any(label and "OVER BUDGET" in str(label) for label in gameplan_labels):
        problems.append("Gameplan sheet missing the over-budget row")
    return problems


def main() -> int:
    tests = [
        test_happy_swap_path,
        test_unfixable_ships_flagged_never_raises,
        test_model_cannot_touch_compliant_cards,
        test_over_cap_commander_is_flagged_not_swapped,
        test_pricing_unavailable_skips_cleanly,
        test_synergy_recheck_notes_a_dip,
        test_select_stage_commander_price_repick,
        test_email_and_xlsx_budget_callouts,
    ]
    all_problems: dict[str, list[str]] = {}
    for test in tests:
        problems = test()
        if problems:
            all_problems[test.__name__] = problems

    if all_problems:
        for name, problems in all_problems.items():
            print(f"FAILED: {name}", file=sys.stderr)
            for p in problems:
                print(f"  - {p}", file=sys.stderr)
        return 1

    print(f"OK: all {len(tests)} budget-pass checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
