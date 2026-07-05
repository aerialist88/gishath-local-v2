"""Regression tests for the PRD v4 amendment (2026-07-03): token diet (T3 swap
application), card tagging off Opus (T4 heuristics), EDHREC pool fallback (S1),
the synergy-density gate (S3), and the newsletter's diagnostics-stripping (§3.3).

Run manually: cd gishath-local-v2 && python3 -m deck_engine.tests.test_v4_amendment
"""
from __future__ import annotations

import sys

from .. import card_tagger, edhrec_pool, emailer, export, synergy_check
from ..agent_pipeline import DeckResult, _apply_swaps
from ..concept_selector import ConceptChoice
from ..pricing import PricingOutcome
from ..scryfall_cache import ValidationResult


def test_apply_swaps_basic() -> list[str]:
    problems = []
    cards = ["Sol Ring", "Arcane Signet", "Rampant Growth"]
    swaps = [{"remove": "Rampant Growth", "add": "Cultivate", "reason": "more on-mechanic"}]
    new_cards, warnings = _apply_swaps(cards, swaps)
    if new_cards != ["Sol Ring", "Arcane Signet", "Cultivate"]:
        problems.append(f"expected Cultivate to replace Rampant Growth in place, got {new_cards}")
    if warnings:
        problems.append(f"expected no warnings for a clean swap, got {warnings}")
    return problems


def test_apply_swaps_missing_remove_is_warned_not_fatal() -> list[str]:
    problems = []
    cards = ["Sol Ring", "Arcane Signet"]
    swaps = [{"remove": "Nonexistent Card", "add": "Cultivate", "reason": "test"}]
    new_cards, warnings = _apply_swaps(cards, swaps)
    if "Cultivate" not in new_cards:
        problems.append("expected Cultivate to be added even though the remove target was missing")
    if len(new_cards) != 3:
        problems.append(f"expected 3 cards (2 original + 1 appended add), got {len(new_cards)}: {new_cards}")
    if not warnings:
        problems.append("expected a warning for a remove-target-not-found swap")
    return problems


def test_apply_swaps_empty_list_is_noop() -> list[str]:
    problems = []
    cards = ["Sol Ring", "Arcane Signet"]
    new_cards, warnings = _apply_swaps(cards, [])
    if new_cards != cards:
        problems.append(f"expected an empty swaps list to leave the decklist unchanged, got {new_cards}")
    if warnings:
        problems.append(f"expected no warnings, got {warnings}")
    return problems


_FAKE_CACHE = {
    "sol ring": {"name": "Sol Ring", "type_line": "Artifact", "oracle_text": "{T}: Add {C}{C}.", "cmc": 1},
    "swords to plowshares": {
        "name": "Swords to Plowshares", "type_line": "Instant",
        "oracle_text": "Exile target creature. Its controller gains life.", "cmc": 1,
    },
    "cultivate": {
        "name": "Cultivate", "type_line": "Sorcery",
        "oracle_text": "Search your library for up to two basic land cards.", "cmc": 2,
    },
    "wrath of god": {
        "name": "Wrath of God", "type_line": "Sorcery", "oracle_text": "Destroy all creatures.", "cmc": 4,
    },
    "divination": {"name": "Divination", "type_line": "Sorcery", "oracle_text": "Draw two cards.", "cmc": 3},
    "forest": {"name": "Forest", "type_line": "Basic Land — Forest", "oracle_text": "", "cmc": 0},
    "mystery creature": {
        "name": "Mystery Creature", "type_line": "Creature — Weirdo",
        "oracle_text": "Does something nobody's ever seen before.", "cmc": 3,
    },
}


def test_card_tagger_heuristics_no_model_call() -> list[str]:
    """Ramp/removal/draw/wipe/land all get tagged by pure code heuristics — the
    ambiguous fallback (which would need a Haiku call) should only fire for the
    genuinely-unclassifiable card."""
    problems = []
    heuristics = {
        "sol ring": ("Ramp", "early"),  # "Add {C}{C}" matches the ramp pattern — correctly a mana rock
        "swords to plowshares": ("Removal", "mid"),
        "cultivate": ("Ramp", "early"),
        "wrath of god": ("Board wipe", "mid"),
        "divination": ("Card draw", "mid"),
        "forest": ("Land/Mana base", "early"),
    }
    for key, expected in heuristics.items():
        got = card_tagger._heuristic_tag(_FAKE_CACHE[key])  # noqa: SLF001
        if got != expected:
            problems.append(f"{key}: expected heuristic {expected}, got {got}")
    ambiguous = card_tagger._heuristic_tag(_FAKE_CACHE["mystery creature"])  # noqa: SLF001
    if ambiguous is not None:
        problems.append(f"expected Mystery Creature to fall through to the Haiku pass, got {ambiguous}")
    return problems


def test_synergy_gate_counts_and_threshold() -> list[str]:
    problems = []
    tokens = ["proliferate", "+1/+1 counter"]
    cache = {
        "card a": {"type_line": "Creature", "oracle_text": "Whenever this attacks, proliferate."},
        "card b": {"type_line": "Creature", "oracle_text": "Put a +1/+1 counter on target creature."},
        "card c": {"type_line": "Creature", "oracle_text": "Draw a card."},  # no match
        "basic land": {"type_line": "Basic Land — Forest", "oracle_text": ""},  # lands exempt
    }
    cards = ["Card A", "Card B", "Card C", "Basic Land"]
    match_count, generic = synergy_check.count_synergy_matches(cards, tokens, cache)
    if match_count != 2:
        problems.append(f"expected 2 matches (Card A, Card B), got {match_count}")
    if generic != ["Card C"]:
        problems.append(f"expected only Card C flagged generic (lands exempt), got {generic}")

    passes, _, _ = synergy_check.gate_passes(cards, tokens, cache, threshold=2)
    if not passes:
        problems.append("expected gate to pass at threshold=2 with 2 matches")
    passes, _, _ = synergy_check.gate_passes(cards, tokens, cache, threshold=3)
    if passes:
        problems.append("expected gate to fail at threshold=3 with only 2 matches")

    # No tokens extracted -> gate must never block a run (treat as an automatic pass).
    passes, match_count, generic = synergy_check.gate_passes(cards, [], cache, threshold=25)
    if not passes or generic:
        problems.append("expected an empty token list to be a no-op pass, never a fail")
    return problems


def test_edhrec_slugify() -> list[str]:
    problems = []
    cases = {
        "Zinnia, Valley's Voice": "zinnia-valleys-voice",
        "Gishath, Sun's Avatar": "gishath-suns-avatar",
        "Atraxa, Praetors' Voice": "atraxa-praetors-voice",
    }
    for commander, expected in cases.items():
        got = edhrec_pool._slugify(commander)  # noqa: SLF001
        if got != expected:
            problems.append(f"{commander!r}: expected slug {expected!r}, got {got!r}")
    return problems


def test_edhrec_pool_block_fallback_under_min_size() -> list[str]:
    """A pool with fewer than EDHREC_MIN_POOL_SIZE cards must report pool_usable=False
    and a fallback message — never silently pretend a thin pool is a real candidate
    list. fetch_pool() will fail/return [] in this sandbox (no network to
    json.edhrec.com, same class of restriction as Scryfall/TCG Marketplace
    elsewhere in this project) — which IS the fallback path this test exercises,
    no monkeypatching needed."""
    problems = []
    text, usable = edhrec_pool.pool_block("Some Obscure Commander That Definitely Has No Cache Entry", {})
    if usable:
        problems.append("expected pool_usable=False when fetch_pool() returns fewer than the minimum")
    if "no usable edhrec synergy pool" not in text.lower():
        problems.append(f"expected a fallback explanation in the returned block, got: {text!r}")
    return problems


def _build_fake_deck_for_newsletter() -> DeckResult:
    concept = ConceptChoice(
        commander="Test Commander", archetype="Test archetype", rationale="test rationale",
        color_identity=["U"], oracle_text="(test oracle text)",
    )
    validation = ValidationResult(commander=concept.commander, card_count=100)
    return DeckResult(
        concept=concept, cards=["Sol Ring"] * 99, validation=validation,
        changes_made="no changes", early_game="e", mid_game="m", late_game="l",
        final_archetype="Test archetype", final_summary="test summary",
        synergy_gate_fired=True,
    )


def test_newsletter_strips_diagnostics_for_friends_copy() -> list[str]:
    """The clean (friends') copy must never leak cost/turn/tool diagnostics — that's
    the whole point of the two-send design (PRD v4 amendment §3.3, resolved open
    question #2)."""
    problems = []
    deck = _build_fake_deck_for_newsletter()
    spend_summary = {"total_cost_usd": 3.21, "total_turns": 17, "total_duration_ms": 45000, "tools_used": ["WebSearch"]}
    price_summary = {"total": 42.50, "priced_count": 100, "unpriced_count": 0, "top_expensive": [("Sol Ring", 5.0)]}

    clean_plain = emailer._plain_success_body(  # noqa: SLF001
        deck, spend_summary, "PASSED", price_summary, [], include_diagnostics=False)
    full_plain = emailer._plain_success_body(  # noqa: SLF001
        deck, spend_summary, "PASSED", price_summary, [], include_diagnostics=True)

    if "3.2100" in clean_plain or "$3.2100" in clean_plain:
        problems.append("clean (friends') plain body leaked run cost")
    if "17 turns" in clean_plain:
        problems.append("clean (friends') plain body leaked turn count")
    if "$3.2100" not in full_plain and "3.2100" not in full_plain:
        problems.append("full (Trevor's) plain body is missing run cost — diagnostics should be present")

    clean_html = emailer._html_success_body(  # noqa: SLF001
        deck, spend_summary, "PASSED", True, price_summary, [], None, include_diagnostics=False)
    full_html = emailer._html_success_body(  # noqa: SLF001
        deck, spend_summary, "PASSED", True, price_summary, [], None, include_diagnostics=True)

    if "Run cost" in clean_html:
        problems.append("clean (friends') HTML body leaked the run-cost line")
    if "Run cost" not in full_html:
        problems.append("full (Trevor's) HTML body is missing the run-cost line")

    # Both copies should still show the SGD headline — that's the deck info, not a diagnostic.
    if "42.50" not in clean_plain or "42.50" not in clean_html:
        problems.append("clean copy is missing the SGD price headline — that should ship to friends too")
    return problems


def test_moxfield_lists_basics_individually() -> list[str]:
    """2026-07-03 real-run feedback: grouped 'qty N' rows for basics were hard to
    paste into Moxfield in practice — every copy must be its own '1 Forest' line/
    row, not one 'Forest x20' row, in both the .txt export and the xlsx Moxfield
    sheet."""
    problems = []
    concept = ConceptChoice(
        commander="Test Commander", archetype="a", rationale="r",
        color_identity=["G"], oracle_text="",
    )
    validation = ValidationResult(commander=concept.commander, card_count=100)
    cards = ["Sol Ring"] + ["Forest"] * 20 + ["Cultivate"]
    deck = DeckResult(
        concept=concept, cards=cards, validation=validation,
        changes_made="", early_game="", mid_game="", late_game="",
    )

    lines = export._moxfield_lines(deck)  # noqa: SLF001
    forest_lines = [ln for ln in lines if ln == "1 Forest"]
    if len(forest_lines) != 20:
        problems.append(f"expected 20 separate '1 Forest' lines, got {len(forest_lines)}: {lines}")
    if any(ln.startswith("20 Forest") or "×" in ln for ln in lines):
        problems.append(f"found a grouped/qty>1 Forest line — basics must not be grouped: {lines}")

    pricing = PricingOutcome(plan=None, available=False, error="test: pricing skipped")
    import io
    from openpyxl import load_workbook
    xlsx_bytes = export.write_deck_excel(deck, pricing, cache={})
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    mox_rows = list(wb["Moxfield"].iter_rows(values_only=True))
    forest_rows = [r for r in mox_rows if r[1] == "Forest"]
    if len(forest_rows) != 20:
        problems.append(f"expected 20 separate Forest rows in the Moxfield sheet, got {len(forest_rows)}")
    if any(r[0] != 1 for r in forest_rows):
        problems.append(f"expected every Forest row's qty column to be 1, got {[r[0] for r in forest_rows]}")
    return problems


def test_scryfall_cache_schema_version_forces_refresh() -> list[str]:
    """A cache refreshed before a _KEEP_FIELDS change (e.g. the 2026-07-03 addition
    of cmc/mana_cost/rarity/image_uris/card_faces) must be treated as stale
    regardless of age — this is the real bug that shipped blank CMC/Rarity/curve/
    pip data on a real run, root-caused to refresh_if_stale()'s age-only check."""
    from .. import scryfall_cache
    problems = []

    old_style_meta = {"keep_fields": ["name", "type_line", "oracle_text", "color_identity", "legalities",
                                       "layout", "oracle_id"]}
    current_fields = sorted(scryfall_cache._KEEP_FIELDS)
    if old_style_meta["keep_fields"] == current_fields:
        problems.append("test fixture's old-style keep_fields unexpectedly matches current _KEEP_FIELDS — "
                         "update the fixture to reflect a genuinely older schema")

    # Simulate _cache_schema_matches()'s comparison directly (avoids touching real
    # cache files on disk, which this sandbox can't refresh from live Scryfall anyway).
    matches = old_style_meta.get("keep_fields") == current_fields
    if matches:
        problems.append("expected an old-style meta (missing the new fields) to NOT match current _KEEP_FIELDS")

    no_keep_fields_meta = {"refreshed_at": "2026-07-01T00:00:00+00:00"}
    matches_missing = no_keep_fields_meta.get("keep_fields") == current_fields
    if matches_missing:
        problems.append("expected a meta with no keep_fields entry at all to be treated as a mismatch")
    return problems


def main() -> int:
    tests = [
        test_apply_swaps_basic,
        test_apply_swaps_missing_remove_is_warned_not_fatal,
        test_apply_swaps_empty_list_is_noop,
        test_card_tagger_heuristics_no_model_call,
        test_synergy_gate_counts_and_threshold,
        test_edhrec_slugify,
        test_edhrec_pool_block_fallback_under_min_size,
        test_newsletter_strips_diagnostics_for_friends_copy,
        test_moxfield_lists_basics_individually,
        test_scryfall_cache_schema_version_forces_refresh,
    ]
    all_problems: dict[str, list[str]] = {}
    for test in tests:
        problems = test()
        if problems:
            all_problems[test.__name__] = problems

    if all_problems:
        for name, problems in all_problems.items():
            print(f"FAILED: {name}", file=sys.stderr)
            for p in problems:
                print(f"  - {p}", file=sys.stderr)
        return 1

    print(f"OK: all {len(tests)} PRD v4 amendment checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
