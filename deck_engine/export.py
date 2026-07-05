"""
deck_engine/export.py — stage 7: build the .xlsx + Moxfield .txt (PRD §2.1 /
§4d, extended by PRD v4 amendment §3.3).

Sheet 1 "Moxfield" — plain `qty | name` decklist, commander first, one row
per card COPY (basic lands listed individually — "Mountain" x8 is 8 separate
rows, not one qty-8 row) — imports cleanly into Moxfield's text/CSV deck
import, and is easy to eyeball/edit by hand. Reverted 2026-07-03 from an
earlier grouped-basics version: Trevor found the grouped qty rows harder to
paste into Moxfield in practice than one line per copy. save_moxfield_txt()
writes the same list as a plain .txt attachment (PRD v4 amendment §3.3) for
pilots who'd rather paste than open the xlsx.
Sheet 2 "Breakdown" — per card copy: SG price (cheapest, Strategy A from
pricing.deck_price_summary()/cheapest_by_card()), store, role, phase, CMC,
type, rarity. Sorted by role then price (descending) within role, commander
pinned first. Missing prices are flagged, never faked (PRD §2.4).
Sheet 3 "Gameplan" — archetype/summary/early-mid-late text, changes made,
priced total (repeated here per PRD v4 amendment §3.3 so it's visible without
flipping to Breakdown).
Sheet 4 "Stats" — mana curve, colour-pip counts, role counts (PRD v4
amendment §3.3) — computed here from the Scryfall cache, not asked of the
model (this is presentation metadata, same reasoning as T4's card tagging).

Colour palette borrowed from gishath-local-v2/export/excel.py for visual
consistency with the rest of the project, not reused code (the sheet shapes
are different enough — Moxfield export vs. shopping-plan export — that a new
small writer was clearer than bolting onto write_excel()).
"""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config, pricing as pricing_mod
from .agent_pipeline import DeckResult
from .pricing import PricingOutcome

BG_HEADER = "F3F4F6"
FG_HEADER = "6D28D9"
FG_CARD = "111827"
FG_PRICE = "065F46"
FG_MISSING = "B91C1C"
BG_MISSING = "FEF2F2"
FG_STORE = "3730A3"
FG_ROLE = "374151"
FG_PHASE_EARLY = "1D4ED8"
FG_PHASE_MID = "B45309"
FG_PHASE_LATE = "15803D"
_HEADER_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))
_PHASE_COLOR = {"early": FG_PHASE_EARLY, "mid": FG_PHASE_MID, "late": FG_PHASE_LATE}

# Roughly the order Trevor's own gameplan sheet already narrates in (early ->
# mid -> late utility), commander pinned separately at the very top by the
# caller. Anything not in this list sorts after, alphabetically — a missing/
# unrecognized role is cosmetic, never worth failing an export over.
_ROLE_SORT_ORDER = [
    "Land/Mana base", "Ramp", "Card draw", "Removal", "Board wipe",
    "Protection", "Synergy piece", "Win condition", "Interaction",
]
_MANA_SYMBOL_RE = re.compile(r"\{([^}]+)\}")


def _style_header_row(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color=FG_HEADER, size=11)
        cell.fill = PatternFill("solid", fgColor=BG_HEADER)
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(vertical="center")


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _role_sort_key(role: str) -> tuple[int, str]:
    try:
        return _ROLE_SORT_ORDER.index(role), ""
    except ValueError:
        return len(_ROLE_SORT_ORDER), role or "~"  # unrecognized/blank roles sort last, alphabetically among themselves


def _build_moxfield_sheet(wb: Workbook, deck: DeckResult) -> None:
    ws = wb.active
    ws.title = "Moxfield"
    _style_header_row(ws, 1, ["qty", "name"])

    ws.cell(row=2, column=1, value=1)
    c = ws.cell(row=2, column=2, value=deck.concept.commander)
    c.font = Font(bold=True, color=FG_CARD)

    for i, card in enumerate(deck.cards, start=3):
        ws.cell(row=i, column=1, value=1)
        ws.cell(row=i, column=2, value=card).font = Font(color=FG_CARD)

    _autosize(ws, {1: 6, 2: 42})
    ws.freeze_panes = "A2"


def _moxfield_lines(deck: DeckResult) -> list[str]:
    lines = [f"1 {deck.concept.commander}"]
    lines.extend(f"1 {card}" for card in deck.cards)
    return lines


def write_moxfield_txt(deck: DeckResult) -> bytes:
    """Plain Moxfield-importable decklist text (PRD v4 amendment §3.3) — same
    qty/name grouping as the Moxfield sheet, as a standalone attachment for
    pilots who'd rather paste text than open the xlsx."""
    return ("\n".join(_moxfield_lines(deck)) + "\n").encode("utf-8")


def save_moxfield_txt(deck: DeckResult, run_id: str) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M")
    safe_commander = "".join(c if c.isalnum() or c in " -_" else "" for c in deck.concept.commander).strip()
    filename = f"{timestamp}_{safe_commander}_{run_id[:8]}_moxfield.txt"
    path = config.OUTPUT_DIR / filename
    path.write_bytes(write_moxfield_txt(deck))
    return path


def _build_breakdown_sheet(wb: Workbook, deck: DeckResult, pricing: PricingOutcome, cache: dict) -> None:
    ws = wb.create_sheet("Breakdown")
    headers = ["Card", "SG Price (cheapest)", "Store", "Role", "Phase", "CMC", "Type", "Rarity"]
    _style_header_row(ws, 1, headers)

    cheapest = pricing_mod.cheapest_by_card(pricing)

    commander = deck.concept.commander

    def _row_data(name: str) -> dict:
        key = name.strip().lower()
        tag = deck.card_tags.get(key, {"role": "", "phase": ""})
        card = cache.get(key) or {}
        price_info = cheapest.get(key)
        return {"name": name, "role": tag.get("role", ""), "phase": tag.get("phase", ""),
                "card": card, "price_info": price_info}

    commander_row = _row_data(commander)
    other_rows = [_row_data(name) for name in deck.cards]
    other_rows.sort(key=lambda r: (_role_sort_key(r["role"]), -(r["price_info"][0] if r["price_info"] else 0)))

    total_price = 0.0
    priced_count = 0
    unpriced_count = 0
    row = 2

    for i, data in enumerate([commander_row] + other_rows):
        name, role, phase, card = data["name"], data["role"], data["phase"], data["card"]
        if i == 0:
            role = "Commander"  # pinned display override — matches agent_pipeline.run_pipeline()'s own override

        name_cell = ws.cell(row=row, column=1, value=name)
        name_cell.font = Font(color=FG_CARD, bold=(i == 0))

        price_cell = ws.cell(row=row, column=2)
        store_cell = ws.cell(row=row, column=3)
        price_info = data["price_info"]
        if price_info is not None:
            price_val, store = price_info
            price_cell.value = price_val
            price_cell.number_format = '"SGD" #,##0.00'
            price_cell.font = Font(color=FG_PRICE)
            store_cell.value = store
            store_cell.font = Font(color=FG_STORE)
            total_price += price_val
            priced_count += 1
        else:
            price_cell.value = "unavailable"
            price_cell.font = Font(color=FG_MISSING, italic=True)
            price_cell.fill = PatternFill("solid", fgColor=BG_MISSING)
            store_cell.value = "—"
            unpriced_count += 1

        ws.cell(row=row, column=4, value=role).font = Font(color=FG_ROLE)
        phase_cell = ws.cell(row=row, column=5, value=phase)
        phase_cell.font = Font(color=_PHASE_COLOR.get(phase, FG_ROLE), bold=bool(phase))

        ws.cell(row=row, column=6, value=card.get("cmc")).font = Font(color=FG_ROLE)
        ws.cell(row=row, column=7, value=card.get("type_line", "")).font = Font(color=FG_ROLE)
        ws.cell(row=row, column=8, value=(card.get("rarity") or "").title()).font = Font(color=FG_ROLE)

        row += 1

    # Total row — sums only priced cards, clearly labelled if some are missing so
    # the total never silently understates the real cost (PRD §2.4).
    total_border = Border(top=Side(style="thin", color=FG_HEADER))
    label = "Total (priced cards only)" if unpriced_count else "Total"
    if unpriced_count:
        label += f" — {unpriced_count} card(s) unpriced, excluded"
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = Font(bold=True, color=FG_HEADER)
    label_cell.border = total_border

    total_cell = ws.cell(row=row, column=2, value=round(total_price, 2))
    total_cell.number_format = '"SGD" #,##0.00'
    total_cell.font = Font(bold=True, color=FG_PRICE)
    total_cell.border = total_border

    for col in (3, 4, 5, 6, 7, 8):
        ws.cell(row=row, column=col).border = total_border

    row += 1
    if not pricing.available:
        note_row = row + 1
        ws.cell(row=note_row, column=1,
                value=f"Pricing unavailable this run: {pricing.error}").font = Font(color=FG_MISSING, italic=True)

    _autosize(ws, {1: 30, 2: 16, 3: 22, 4: 18, 5: 8, 6: 6, 7: 24, 8: 12})
    ws.freeze_panes = "A2"


def _budget_rows(budget) -> list[tuple[str, str]]:
    """Gameplan-sheet rows for the §3.4 budget pass — swaps listed distinctly from
    optimize's changes, over-cap leftovers flagged loudly, nothing shown at all if
    the pass did nothing (no noise on a clean run)."""
    if budget is None:
        return []
    rows: list[tuple[str, str]] = []
    if budget.swaps_made:
        lines = []
        for removed, removed_price, added, added_price, reason in budget.swaps_made:
            added_str = f"SGD {added_price:.2f}" if added_price is not None else "unpriced"
            lines.append(f"{removed} (SGD {removed_price:.2f}) → {added} ({added_str}) — {reason}")
        rows.append((f"Budget pass (cap SGD {budget.cap:.0f}/card)", "\n".join(lines)))
    if budget.over_budget:
        over = ", ".join(f"{c} (SGD {p:.2f})" for c, p in budget.over_budget)
        rows.append((
            "OVER BUDGET — shipped flagged",
            f"No suitable budget substitute was found for: {over}. Shipped anyway per the "
            "flag-never-block rule — swap these manually if the price matters.",
        ))
    if budget.synergy_note:
        rows.append(("Synergy after budget swaps", budget.synergy_note))
    return rows


def _build_summary_sheet(wb: Workbook, deck: DeckResult, pricing: PricingOutcome, budget=None) -> None:
    ws = wb.create_sheet("Gameplan")
    _style_header_row(ws, 1, ["Section", "Notes"])

    all_cards = [deck.concept.commander] + deck.cards
    price_summary = pricing_mod.deck_price_summary(pricing, all_cards)
    total_line = f"SGD {price_summary['total']:.2f}"
    if price_summary["unpriced_count"]:
        total_line += f" ({price_summary['unpriced_count']} card(s) unpriced, excluded)"

    rows = [
        ("Commander", deck.concept.commander),
        ("Archetype", deck.final_archetype),
        ("Why this pick tonight", deck.final_summary),
        ("Deck total (SGD)", total_line),
        ("Early game", deck.early_game),
        ("Mid game", deck.mid_game),
        ("Late game", deck.late_game),
        ("Changes made during optimize pass", deck.changes_made or "no changes"),
    ]
    rows.extend(_budget_rows(budget))
    if not deck.edhrec_pool_used:
        rows.append((
            "EDHREC pool",
            "No usable synergy pool found for this commander tonight (too new/obscure, or the "
            "endpoint was unreachable) — built without it, same as before this feature.",
        ))
    if deck.synergy_gate_fired:
        rows.append((
            "Synergy gate",
            "The code-level synergy-density check flagged this deck as under-leaning on the "
            "commander's specific mechanic and triggered a targeted repair pass — the version "
            "here is the corrected result.",
        ))
    if deck.retried:
        rows.append((
            "Rebuilt mid-run",
            f"The first build was discarded — its gameplan depended on an ability the commander/a key "
            f"card doesn't actually have: \"{deck.retry_reason}\". Re-ran ideation with the real oracle "
            f"text and a note about that specific mistake; this deck is the corrected result.",
        ))
    rows.append(("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")))
    for i, (label, text) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, color=FG_HEADER)
        cell = ws.cell(row=i, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autosize(ws, {1: 26, 2: 90})


def _compute_stats(deck: DeckResult, cache: dict) -> dict:
    """Mana curve, colour-pip counts, and role tallies — computed here from the
    Scryfall cache rather than asked of the model (PRD v4 amendment §3.3; this is
    presentation metadata, same reasoning as T4's card tagging). Iterates every
    card COPY individually (a repeated basic land counts once per copy) — same
    numbers either way, just no longer routed through display grouping."""
    all_cards = [deck.concept.commander] + list(deck.cards)
    curve = {"0": 0, "1": 0, "2": 0, "3": 0, "4": 0, "5": 0, "6+": 0}
    pips = {"W": 0, "U": 0, "B": 0, "R": 0, "G": 0, "C": 0}
    role_counts: dict[str, int] = {}

    for name in all_cards:
        key = name.strip().lower()
        card = cache.get(key)
        tag = deck.card_tags.get(key, {})
        role = tag.get("role") or "(untagged)"
        role_counts[role] = role_counts.get(role, 0) + 1

        if card is None:
            continue
        type_line = (card.get("type_line") or "").lower()
        is_land = "land" in type_line

        if not is_land:
            cmc = card.get("cmc")
            if cmc is not None:
                bucket = "6+" if cmc >= 6 else str(int(cmc))
                curve[bucket] = curve.get(bucket, 0) + 1

        faces = card.get("card_faces") or None
        mana_costs = (
            [f.get("mana_cost", "") for f in faces if isinstance(f, dict)]
            if faces else [card.get("mana_cost") or ""]
        )
        for mana_cost in mana_costs:
            for symbol in _MANA_SYMBOL_RE.findall(mana_cost or ""):
                symbol_upper = symbol.upper()
                for c in "WUBRG":
                    if c in symbol_upper:
                        pips[c] += 1
                if symbol_upper == "C":
                    pips["C"] += 1

    return {"curve": curve, "pips": pips, "role_counts": role_counts}


def _build_stats_sheet(wb: Workbook, deck: DeckResult, cache: dict) -> None:
    ws = wb.create_sheet("Stats")
    stats = _compute_stats(deck, cache)

    row = 1
    _style_header_row(ws, row, ["Mana Curve (nonland)", "Count"])
    row += 1
    for bucket in ["0", "1", "2", "3", "4", "5", "6+"]:
        ws.cell(row=row, column=1, value=f"CMC {bucket}").font = Font(color=FG_CARD)
        ws.cell(row=row, column=2, value=stats["curve"].get(bucket, 0)).font = Font(color=FG_CARD)
        row += 1

    row += 1
    _style_header_row(ws, row, ["Colour Pips", "Count"])
    row += 1
    pip_labels = {"W": "White", "U": "Blue", "B": "Black", "R": "Red", "G": "Green", "C": "Colourless"}
    for symbol, label in pip_labels.items():
        ws.cell(row=row, column=1, value=label).font = Font(color=FG_CARD)
        ws.cell(row=row, column=2, value=stats["pips"].get(symbol, 0)).font = Font(color=FG_CARD)
        row += 1

    row += 1
    _style_header_row(ws, row, ["Role", "Count"])
    row += 1
    for role, count in sorted(stats["role_counts"].items(), key=lambda kv: (-kv[1], kv[0])):
        ws.cell(row=row, column=1, value=role).font = Font(color=FG_CARD)
        ws.cell(row=row, column=2, value=count).font = Font(color=FG_CARD)
        row += 1

    _autosize(ws, {1: 24, 2: 10})


def write_deck_excel(deck: DeckResult, pricing: PricingOutcome, cache: dict | None = None, budget=None) -> bytes:
    """Builds the full workbook (Moxfield + Breakdown + Gameplan + Stats sheets)
    and returns raw bytes. `cache` powers the CMC/type/rarity columns and Stats
    sheet (PRD v4 amendment §3.3) — pass the same Scryfall cache dict the pipeline
    already loaded; an empty dict degrades those columns/sheet to blanks/zeros
    rather than raising (useful for tests that don't need real card data)."""
    cache = cache if cache is not None else {}
    wb = Workbook()
    _build_moxfield_sheet(wb, deck)
    _build_breakdown_sheet(wb, deck, pricing, cache)
    _build_summary_sheet(wb, deck, pricing, budget=budget)
    _build_stats_sheet(wb, deck, cache)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_deck_excel(
    deck: DeckResult, pricing: PricingOutcome, run_id: str, cache: dict | None = None, budget=None,
) -> Path:
    """Writes the workbook to config.OUTPUT_DIR and returns its path."""
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M")
    safe_commander = "".join(c if c.isalnum() or c in " -_" else "" for c in deck.concept.commander).strip()
    filename = f"{timestamp}_{safe_commander}_{run_id[:8]}.xlsx"
    path = config.OUTPUT_DIR / filename
    path.write_bytes(write_deck_excel(deck, pricing, cache, budget=budget))
    return path


def deck_json_payload(
    deck: DeckResult, pricing: PricingOutcome, run_id: str, cache: dict | None = None,
    budget=None, spend_summary: dict | None = None,
    xlsx_path: Path | None = None, moxfield_txt_path: Path | None = None,
) -> dict:
    """Machine-readable record of a finished deck — everything the Atelier UI's
    deck view needs (decklist by role with prices, curve, gameplan, run
    diagnostics), the same data the xlsx sheets present, in one JSON document.
    Written alongside the xlsx by run.py so past commissions stay browsable
    without re-parsing spreadsheets."""
    cache = cache if cache is not None else {}
    cheapest = pricing_mod.cheapest_by_card(pricing)
    all_cards = [deck.concept.commander] + list(deck.cards)
    price_summary = pricing_mod.deck_price_summary(pricing, all_cards)

    cards = []
    for i, name in enumerate(all_cards):
        key = name.strip().lower()
        tag = deck.card_tags.get(key, {})
        card = cache.get(key) or {}
        price_info = cheapest.get(key)
        price = round(price_info[0], 2) if price_info else None
        cards.append({
            "name": name,
            "is_commander": i == 0,
            "role": "Commander" if i == 0 else (tag.get("role") or ""),
            "phase": tag.get("phase") or "",
            "price_sgd": price,
            "store": price_info[1] if price_info else None,
            "over_cap": bool(price is not None and price > config.MAX_CARD_PRICE_SGD),
            "cmc": card.get("cmc"),
            "type_line": card.get("type_line") or "",
            "rarity": (card.get("rarity") or ""),
        })

    v = deck.validation
    return {
        "schema": 1,
        "run_id": run_id,
        "run_id8": run_id[:8],
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "commander": deck.concept.commander,
        "archetype": deck.final_archetype or deck.concept.archetype,
        "summary": deck.final_summary,
        "colors": deck.concept.color_identity,
        "bracket": config.BRACKET,
        "legal": bool(getattr(v, "is_valid", False)),
        "synergy_gate_fired": deck.synergy_gate_fired,
        "edhrec_pool_used": deck.edhrec_pool_used,
        "retried": deck.retried,
        "retry_reason": deck.retry_reason,
        "cards": cards,
        "price": {
            "available": pricing.available,
            "total_sgd": round(price_summary["total"], 2),
            "priced_count": price_summary["priced_count"],
            "unpriced_count": price_summary["unpriced_count"],
            "top_expensive": [[n, round(p, 2)] for n, p in price_summary["top_expensive"]],
            "per_card_cap_sgd": config.MAX_CARD_PRICE_SGD,
            "over_budget": [[c, round(p, 2)] for c, p in (budget.over_budget if budget else [])],
            "swaps_made": len(budget.swaps_made) if budget else 0,
        },
        "gameplan": {
            "early": deck.early_game,
            "mid": deck.mid_game,
            "late": deck.late_game,
            "changes_made": deck.changes_made,
        },
        "stats": _compute_stats(deck, cache),
        "spend": spend_summary or {},
        "files": {
            "xlsx": Path(xlsx_path).name if xlsx_path else None,
            "moxfield_txt": Path(moxfield_txt_path).name if moxfield_txt_path else None,
        },
    }


def save_deck_json(
    deck: DeckResult, pricing: PricingOutcome, run_id: str, cache: dict | None = None,
    budget=None, spend_summary: dict | None = None,
    xlsx_path: Path | None = None, moxfield_txt_path: Path | None = None,
) -> Path:
    """Writes the deck JSON record to config.OUTPUT_DIR and returns its path."""
    import json as _json
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M")
    safe_commander = "".join(c if c.isalnum() or c in " -_" else "" for c in deck.concept.commander).strip()
    path = config.OUTPUT_DIR / f"{timestamp}_{safe_commander}_{run_id[:8]}_deck.json"
    payload = deck_json_payload(deck, pricing, run_id, cache=cache, budget=budget,
                                spend_summary=spend_summary,
                                xlsx_path=xlsx_path, moxfield_txt_path=moxfield_txt_path)
    path.write_text(_json.dumps(payload, indent=1))
    return path
