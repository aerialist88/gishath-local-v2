"""atelier/archive.py — the gallery: every delivered commission as one record.

New runs write a `*_deck.json` next to their xlsx (export.save_deck_json).
Runs from before the Atelier existed only have the xlsx — those are backfilled
here by reading the workbook's own sheets (Breakdown/Gameplan/Stats carry
everything the deck view needs) and caching the result as the same
`*_deck.json` shape, so the parse happens once per old deck, ever.

Output filenames are `YYYY-MM-DD_HH-MM_<Commander>_<runid8>[ _moxfield.txt | .xlsx | _deck.json ]`.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from deck_engine import config, run_log

_FILENAME_RE = re.compile(
    r"^(?P<ts>\d{4}-\d{2}-\d{2}_\d{2}-\d{2})_(?P<commander>.+)_(?P<id8>[0-9a-f]{8})"
    r"(?P<suffix>_deck\.json|_moxfield\.txt|\.xlsx)$"
)


def _scan() -> dict[str, dict]:
    """id8 -> {ts, commander, xlsx, txt, deck_json} from the output directory."""
    found: dict[str, dict] = {}
    for path in config.OUTPUT_DIR.iterdir():
        m = _FILENAME_RE.match(path.name)
        if not m:
            continue
        rec = found.setdefault(m["id8"], {"ts": m["ts"], "commander": m["commander"], "id8": m["id8"]})
        rec["ts"] = max(rec["ts"], m["ts"])
        if m["suffix"] == ".xlsx":
            rec["xlsx"] = path
        elif m["suffix"] == "_moxfield.txt":
            rec["txt"] = path
        else:
            rec["deck_json"] = path
    return found


def _colors_from_run_log(commander: str) -> list[str]:
    for record in run_log.load_records():
        if record.commander.lower() == commander.lower() and record.colors:
            return record.colors
    return []


def _spend_for(id8: str) -> dict:
    """Best-effort spend rollup for a backfilled deck — spend_log keys on the
    full run_id, the filename only keeps 8 chars, so match by prefix."""
    totals = {"total_cost_usd": 0.0, "total_turns": 0}
    try:
        for line in config.SPEND_LOG_PATH.read_text().splitlines():
            if not line.strip():
                continue
            entry = json.loads(line)
            if str(entry.get("run_id", "")).startswith(id8):
                totals["total_cost_usd"] += entry.get("cost_usd", 0.0)
                totals["total_turns"] += entry.get("num_turns", 0)
        totals["total_cost_usd"] = round(totals["total_cost_usd"], 4)
    except Exception:  # noqa: BLE001 — spend is diagnostics, never load-bearing
        pass
    return totals


def _backfill_from_xlsx(rec: dict) -> dict | None:
    """Parse one pre-Atelier xlsx into the deck-JSON shape and cache it."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(rec["xlsx"], read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — a corrupt workbook shouldn't hide the rest of the gallery
        return None

    cards: list[dict] = []
    gameplan: dict = {}
    meta: dict = {}
    stats = {"curve": {}, "pips": {}, "role_counts": {}}

    def _cell(row: tuple, i: int):
        # read_only worksheets trim trailing empty cells, so index defensively
        return row[i] if i < len(row) else None

    if "Breakdown" in wb.sheetnames:
        for i, row in enumerate(wb["Breakdown"].iter_rows(min_row=2, values_only=True)):
            name = _cell(row, 0)
            if not name or str(name).startswith("Total"):
                break
            raw_price = _cell(row, 1)
            price = raw_price if isinstance(raw_price, (int, float)) else None
            store = _cell(row, 2)
            # CK Price (US) column — added after this backfill path was first
            # written, so older exports simply won't have it; index-9 stays
            # None for those rows (row[i] guarded by _cell's length check).
            raw_ck = _cell(row, 8)
            ck_price = raw_ck if isinstance(raw_ck, (int, float)) else None
            cards.append({
                "name": str(name),
                "is_commander": i == 0,
                "role": str(_cell(row, 3) or ""),
                "phase": str(_cell(row, 4) or ""),
                "price_sgd": round(float(price), 2) if price is not None else None,
                "store": str(store) if store and store != "—" else None,
                "over_cap": bool(price is not None and float(price) > config.MAX_CARD_PRICE_SGD),
                "ck_price_usd": round(float(ck_price), 2) if ck_price is not None else None,
                "ck_url": None,  # hyperlinks aren't recoverable from values_only=True reads
                "cmc": _cell(row, 5),
                "type_line": str(_cell(row, 6) or ""),
                "rarity": str(_cell(row, 7) or ""),
            })

    if "Gameplan" in wb.sheetnames:
        plan_rows = {str(r[0] or ""): str(_cell(r, 1) or "")
                     for r in wb["Gameplan"].iter_rows(min_row=2, values_only=True) if r and _cell(r, 0)}
        meta["archetype"] = plan_rows.get("Archetype", "")
        meta["summary"] = plan_rows.get("Why this pick tonight", "")
        gameplan = {
            "early": plan_rows.get("Early game", ""),
            "mid": plan_rows.get("Mid game", ""),
            "late": plan_rows.get("Late game", ""),
            "changes_made": plan_rows.get("Changes made during optimize pass", ""),
        }
        total_match = re.search(r"SGD\s*([\d,]+\.?\d*)", plan_rows.get("Deck total (SGD)", ""))
        if total_match:
            meta["total_sgd"] = float(total_match.group(1).replace(",", ""))

    if "Stats" in wb.sheetnames:
        section = None
        for row in wb["Stats"].iter_rows(values_only=True):
            label = str(row[0] or "")
            if not label:
                continue
            if label.startswith("Mana Curve"):
                section = "curve"
                continue
            if label.startswith("Colour Pips"):
                section = "pips"
                continue
            if label == "Role":
                section = "roles"
                continue
            value = row[1] if len(row) > 1 else None
            if not isinstance(value, (int, float)):
                continue
            if section == "curve" and label.startswith("CMC "):
                stats["curve"][label[4:]] = int(value)
            elif section == "pips":
                symbol = {"White": "W", "Blue": "U", "Black": "B", "Red": "R",
                          "Green": "G", "Colourless": "C"}.get(label)
                if symbol:
                    stats["pips"][symbol] = int(value)
            elif section == "roles":
                stats["role_counts"][label] = int(value)

    wb.close()
    if not cards:
        return None

    commander = cards[0]["name"]
    priced = [c["price_sgd"] for c in cards if c["price_sgd"] is not None]
    top = sorted((c for c in cards if c["price_sgd"] is not None),
                 key=lambda c: -c["price_sgd"])[:5]
    payload = {
        "schema": 1,
        "run_id": rec["id8"],       # full id lost to history; prefix is all we have
        "run_id8": rec["id8"],
        "generated_utc": rec["ts"][:10] + "T" + rec["ts"][11:].replace("-", ":") + ":00+00:00",
        "commander": commander,
        "archetype": meta.get("archetype", ""),
        "summary": meta.get("summary", ""),
        "colors": _colors_from_run_log(commander),
        "bracket": config.BRACKET,
        "legal": True,               # delivered decks passed the validation gate by definition
        "synergy_gate_fired": False,
        "edhrec_pool_used": True,
        "retried": False,
        "retry_reason": "",
        "cards": cards,
        "price": {
            "available": bool(priced),
            "total_sgd": round(meta.get("total_sgd", sum(priced)), 2),
            "priced_count": len(priced),
            "unpriced_count": len(cards) - len(priced),
            "top_expensive": [[c["name"], c["price_sgd"]] for c in top],
            "per_card_cap_sgd": config.MAX_CARD_PRICE_SGD,
            "over_budget": [],
            "swaps_made": 0,
        },
        "gameplan": gameplan,
        "stats": stats,
        "spend": _spend_for(rec["id8"]),
        "files": {
            "xlsx": rec["xlsx"].name,
            "moxfield_txt": rec["txt"].name if rec.get("txt") else None,
        },
        "backfilled_from_xlsx": True,
    }
    cache_path = rec["xlsx"].with_name(rec["xlsx"].name[:-len(".xlsx")] + "_deck.json")
    try:
        cache_path.write_text(json.dumps(payload, indent=1))
    except Exception:  # noqa: BLE001 — read-only disk shouldn't break browsing
        pass
    return payload


def _load_record(rec: dict) -> dict | None:
    if rec.get("deck_json"):
        try:
            return json.loads(rec["deck_json"].read_text())
        except Exception:  # noqa: BLE001 — fall through to backfill
            pass
    if rec.get("xlsx"):
        return _backfill_from_xlsx(rec)
    return None


def list_decks() -> list[dict]:
    """Newest-first gallery summaries. Skips moxfield-txt-only stragglers
    (aborted/test runs with no workbook and no JSON)."""
    out = []
    for rec in _scan().values():
        if not (rec.get("deck_json") or rec.get("xlsx")):
            continue
        if "fake commander" in rec["commander"].lower():
            continue  # sandbox test artifacts (see deck_engine/README pre-flight checklist)
        deck = _load_record(rec)
        if deck is None:
            continue
        out.append({
            "id": deck.get("run_id8", rec["id8"]),
            "ts": rec["ts"],
            "commander": deck.get("commander", rec["commander"]),
            "archetype": deck.get("archetype", ""),
            "colors": deck.get("colors", []),
            "total_sgd": (deck.get("price") or {}).get("total_sgd"),
            "legal": deck.get("legal", True),
            "owner_deck": bool(deck.get("owner_deck")),  # 3vor's own uploads vs guild commissions
        })
    out.sort(key=lambda d: d["ts"], reverse=True)
    return out


def get_deck(id8: str) -> dict | None:
    rec = _scan().get(id8)
    return _load_record(rec) if rec else None


def file_path(id8: str, kind: str) -> Path | None:
    """Path to a deck's downloadable artifact: kind in {xlsx, txt}."""
    rec = _scan().get(id8)
    if not rec:
        return None
    return rec.get("xlsx") if kind == "xlsx" else rec.get("txt") if kind == "txt" else None
